# -*- coding: utf-8 -*-
"""
API برای ژورنال فعالیت‌ها و گزارشات پروژه
ثبت تمام فعالیت‌های مدل‌های AI در هر پروژه
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import Column, String, Text, DateTime, Integer, Boolean, desc, asc
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
import json
import uuid
import asyncio
import logging

from ...core.database import get_db, Base, engine
from ...models.project import Project
from ...services.prompt_helper import PromptHelper

router = APIRouter()
logger = logging.getLogger(__name__)


# ===================== توابع کمکی مدیریت مدل‌ها =====================

def get_active_models(db: Session, preferred_models: List[str] = None) -> List[str]:
    """
    🆕 دریافت لیست مدل‌های فعال
    اگر مدل‌های ترجیحی داده شوند، فقط آنها را برمی‌گرداند (در صورت فعال بودن)
    """
    from ...models.ai_profile import ModelSettings
    from ...core.models_registry import MODEL_REGISTRY

    active_models = []

    # دریافت تنظیمات از دیتابیس
    db_settings = {}
    try:
        settings = db.query(ModelSettings).all()
        db_settings = {s.model_id: s.enabled for s in settings}
    except:
        pass

    for model_id, model in MODEL_REGISTRY.items():
        # چک کردن فعال بودن
        is_enabled = db_settings.get(model_id, model.enabled)
        if is_enabled:
            active_models.append(model_id)

    # اگر مدل‌های ترجیحی داده شدند
    if preferred_models:
        filtered = [m for m in preferred_models if m in active_models]
        if filtered:
            return filtered

    return active_models


def get_replacement_model(db: Session, disabled_model: str, task_type: str = "general") -> str:
    """
    🆕 دریافت مدل جایگزین برای مدل غیرفعال

    الگوریتم:
    1. مدل‌های فعال با امتیاز بالاتر در همان task_type
    2. مدل‌های فعال با همان provider
    3. هر مدل فعال دیگر
    """
    from ...models.ai_profile import ModelSettings, AIProfile
    from ...core.models_registry import MODEL_REGISTRY, get_model

    active_models = get_active_models(db)
    if not active_models:
        raise HTTPException(status_code=503, detail="هیچ مدل فعالی موجود نیست")

    # اگر مدل اصلی فعال است، همان را برگردان
    if disabled_model in active_models:
        return disabled_model

    # پیدا کردن provider مدل غیرفعال
    original_model = get_model(disabled_model)
    original_provider = original_model.provider.value if original_model else None

    # دریافت پروفایل‌ها برای رتبه‌بندی
    profiles = {}
    try:
        profile_records = db.query(AIProfile).filter(AIProfile.model_id.in_(active_models)).all()
        for p in profile_records:
            profiles[p.model_id] = {
                "overall_score": p.overall_score,
                "last_scores": p.last_scores_by_task or {}
            }
    except:
        pass

    # رتبه‌بندی مدل‌های فعال
    candidates = []
    for model_id in active_models:
        model = get_model(model_id)
        if not model:
            continue

        score = profiles.get(model_id, {}).get("overall_score", 50)
        task_score = profiles.get(model_id, {}).get("last_scores", {}).get(task_type, {}).get("overall", 50)

        # امتیاز اضافی برای همان provider
        provider_bonus = 10 if model.provider.value == original_provider else 0

        total_score = (score + task_score) / 2 + provider_bonus
        candidates.append((model_id, total_score))

    # مرتب‌سازی بر اساس امتیاز
    candidates.sort(key=lambda x: x[1], reverse=True)

    replacement = candidates[0][0] if candidates else "claude"
    logger.info(f"[Model Replacement] {disabled_model} -> {replacement}")

    return replacement


def ensure_active_model(db: Session, model_id: str, task_type: str = "general") -> tuple:
    """
    🆕 اطمینان از فعال بودن مدل، در غیر این صورت جایگزینی

    Returns:
        tuple: (final_model_id, was_replaced, replacement_note)
    """
    active_models = get_active_models(db)

    if model_id in active_models:
        return (model_id, False, None)

    replacement = get_replacement_model(db, model_id, task_type)
    note = f"مدل {model_id} غیرفعال بود و با {replacement} جایگزین شد"

    return (replacement, True, note)


# ===================== مدل دیتابیس =====================

class ActivityLog(Base):
    """لاگ فعالیت‌های AI در پروژه"""
    __tablename__ = "activity_logs"

    id = Column(String(50), primary_key=True)
    project_id = Column(String(50), nullable=False, index=True)

    # اطلاعات مدل
    model_id = Column(String(100), nullable=False)
    model_provider = Column(String(50))

    # نوع فعالیت
    activity_type = Column(String(50), nullable=False)  # chat, trigger, analysis, generation

    # محتوا
    prompt = Column(Text)
    response = Column(Text)

    # متادیتا
    tokens_used = Column(Integer, default=0)
    latency_ms = Column(Integer, default=0)
    success = Column(Boolean, default=True)
    error_message = Column(Text)

    # فیلد مرتبط (برای trigger ها)
    field_id = Column(String(50))
    field_name = Column(String(200))

    # تاریخ
    created_at = Column(DateTime, default=datetime.utcnow)

    # اطلاعات اضافی (JSON)
    extra_data = Column(Text)


class DetailedOperation(Base):
    """
    🆕 عملیات جزئی - ثبت سطر به سطر با قابلیت کلیک
    هر ردیف شامل یک عملیات مجزا است که با کلیک، جزئیات کامل نمایش داده می‌شود
    """
    __tablename__ = "detailed_operations"

    id = Column(String(50), primary_key=True)
    project_id = Column(String(50), nullable=False, index=True)
    parent_log_id = Column(String(50), index=True)  # لینک به ActivityLog والد

    # شماره ردیف در گروه عملیات
    sequence_number = Column(Integer, default=0)

    # نوع عملیات
    operation_type = Column(String(100), nullable=False)  # field_create, field_archive, field_merge, memory_update, etc.

    # خلاصه یک خطی (نمایش در لیست)
    summary = Column(String(500), nullable=False)

    # جزئیات کامل (نمایش با کلیک)
    details = Column(Text)  # JSON با تمام جزئیات

    # مقادیر قبل و بعد (برای ردیابی تغییرات)
    before_value = Column(Text)
    after_value = Column(Text)

    # فایل/فیلد مرتبط
    target_type = Column(String(50))  # field, file, memory, config
    target_id = Column(String(100))
    target_name = Column(String(300))

    # وضعیت
    status = Column(String(20), default="completed")  # completed, pending, failed, skipped

    # تاریخ
    created_at = Column(DateTime, default=datetime.utcnow)


class Report(Base):
    """گزارشات تولید شده"""
    __tablename__ = "project_reports"

    id = Column(String(50), primary_key=True)
    project_id = Column(String(50), nullable=False, index=True)

    # نوع گزارش
    report_type = Column(String(50), default="daily")  # daily, weekly, custom

    # محتوا
    title = Column(String(500))
    content = Column(Text)
    summary = Column(Text)

    # آمار
    total_activities = Column(Integer, default=0)
    total_tokens = Column(Integer, default=0)
    models_used = Column(Text)  # JSON list

    # بازه زمانی
    period_start = Column(DateTime)
    period_end = Column(DateTime)

    # تاریخ
    created_at = Column(DateTime, default=datetime.utcnow)
    generated_by = Column(String(100))  # model that generated the report


class ReportTrigger(Base):
    """تنظیمات تریگر گزارش‌گیری"""
    __tablename__ = "report_triggers"

    id = Column(String(50), primary_key=True)
    project_id = Column(String(50), nullable=False, unique=True)

    enabled = Column(Boolean, default=False)
    interval_minutes = Column(Integer, default=1440)  # روزانه
    interval_type = Column(String(20), default="days")

    last_run = Column(DateTime)
    next_run = Column(DateTime)

    # مدل برای تولید گزارش
    report_model = Column(String(100), default="openai")


# ایجاد جداول
Base.metadata.create_all(bind=engine)


# ===================== مدل‌های Pydantic =====================

class ActivityLogCreate(BaseModel):
    """ایجاد لاگ فعالیت"""
    model_id: str
    model_provider: Optional[str] = None
    activity_type: str
    prompt: Optional[str] = None
    response: Optional[str] = None
    tokens_used: int = 0
    latency_ms: int = 0
    success: bool = True
    error_message: Optional[str] = None
    field_id: Optional[str] = None
    field_name: Optional[str] = None
    extra_data: Optional[Dict] = None


class ActivityLogResponse(BaseModel):
    """پاسخ لاگ فعالیت"""
    id: str
    project_id: str
    model_id: str
    model_provider: Optional[str]
    activity_type: str
    prompt: Optional[str]
    response: Optional[str]
    tokens_used: int
    latency_ms: int
    success: bool
    error_message: Optional[str]
    field_id: Optional[str]
    field_name: Optional[str]
    created_at: datetime
    extra_data: Optional[Dict]


class ReportTriggerSettings(BaseModel):
    """تنظیمات تریگر گزارش"""
    enabled: bool = False
    interval_minutes: int = 1440
    interval_type: str = "days"
    report_model: str = "openai"


class DetailedOperationResponse(BaseModel):
    """🆕 پاسخ عملیات جزئی"""
    id: str
    project_id: str
    parent_log_id: Optional[str]
    sequence_number: int
    operation_type: str
    summary: str
    details: Optional[Dict]
    before_value: Optional[str]
    after_value: Optional[str]
    target_type: Optional[str]
    target_id: Optional[str]
    target_name: Optional[str]
    status: str
    created_at: datetime


# ===================== توابع کمکی برای ثبت عملیات سطر به سطر =====================

def log_detailed_operation(
    db,
    project_id: str,
    parent_log_id: str,
    operation_type: str,
    summary: str,
    details: dict = None,
    before_value: str = None,
    after_value: str = None,
    target_type: str = None,
    target_id: str = None,
    target_name: str = None,
    status: str = "completed",
    sequence_number: int = None
) -> DetailedOperation:
    """
    🆕 ثبت یک عملیات جزئی در ژورنال
    این تابع برای ثبت سطر به سطر عملیات استفاده می‌شود
    """
    # اگر sequence_number داده نشده، آخرین شماره را پیدا کن
    if sequence_number is None:
        last_op = db.query(DetailedOperation).filter(
            DetailedOperation.parent_log_id == parent_log_id
        ).order_by(DetailedOperation.sequence_number.desc()).first()
        sequence_number = (last_op.sequence_number + 1) if last_op else 1

    op = DetailedOperation(
        id=f"op_{uuid.uuid4().hex[:12]}",
        project_id=project_id,
        parent_log_id=parent_log_id,
        sequence_number=sequence_number,
        operation_type=operation_type,
        summary=summary,
        details=json.dumps(details, ensure_ascii=False) if details else None,
        before_value=before_value,
        after_value=after_value,
        target_type=target_type,
        target_id=target_id,
        target_name=target_name,
        status=status,
        created_at=datetime.utcnow()
    )
    db.add(op)
    return op


def detailed_op_to_response(op: DetailedOperation) -> Dict:
    """تبدیل عملیات جزئی به دیکشنری"""
    details = {}
    try:
        if op.details:
            details = json.loads(op.details)
    except:
        pass

    return {
        "id": op.id,
        "project_id": op.project_id,
        "parent_log_id": op.parent_log_id,
        "sequence_number": op.sequence_number,
        "operation_type": op.operation_type,
        "summary": op.summary,
        "details": details,
        "before_value": op.before_value[:500] if op.before_value and len(op.before_value) > 500 else op.before_value,
        "after_value": op.after_value[:500] if op.after_value and len(op.after_value) > 500 else op.after_value,
        "target_type": op.target_type,
        "target_id": op.target_id,
        "target_name": op.target_name,
        "status": op.status,
        "created_at": op.created_at.isoformat() if op.created_at else None,
    }


# ===================== توابع کمکی =====================

def log_to_response(log: ActivityLog) -> Dict:
    """تبدیل لاگ به دیکشنری"""
    extra = {}
    try:
        if log.extra_data:
            extra = json.loads(log.extra_data)
    except:
        pass

    return {
        "id": log.id,
        "project_id": log.project_id,
        "model_id": log.model_id,
        "model_provider": log.model_provider,
        "activity_type": log.activity_type,
        "prompt": log.prompt[:200] + "..." if log.prompt and len(log.prompt) > 200 else log.prompt,
        "response": log.response[:300] + "..." if log.response and len(log.response) > 300 else log.response,
        "tokens_used": log.tokens_used,
        "latency_ms": log.latency_ms,
        "success": log.success,
        "error_message": log.error_message,
        "field_id": log.field_id,
        "field_name": log.field_name,
        "created_at": log.created_at.isoformat() if log.created_at else None,
        "extra_data": extra,
    }


def log_to_full_response(log: ActivityLog) -> Dict:
    """تبدیل لاگ به دیکشنری کامل (بدون truncate)"""
    extra = {}
    try:
        if log.extra_data:
            extra = json.loads(log.extra_data)
    except:
        pass

    return {
        "id": log.id,
        "project_id": log.project_id,
        "model_id": log.model_id,
        "model_provider": log.model_provider,
        "activity_type": log.activity_type,
        "prompt": log.prompt,
        "response": log.response,
        "tokens_used": log.tokens_used,
        "latency_ms": log.latency_ms,
        "success": log.success,
        "error_message": log.error_message,
        "field_id": log.field_id,
        "field_name": log.field_name,
        "created_at": log.created_at.isoformat() if log.created_at else None,
        "extra_data": extra,
    }


# ===================== Endpoints =====================

@router.post("/{project_id}/journal/log")
async def create_activity_log(
    project_id: str,
    log_data: ActivityLogCreate,
    db: Session = Depends(get_db)
):
    """ثبت یک فعالیت جدید در ژورنال"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    log = ActivityLog(
        id=f"log_{uuid.uuid4().hex[:12]}",
        project_id=project_id,
        model_id=log_data.model_id,
        model_provider=log_data.model_provider,
        activity_type=log_data.activity_type,
        prompt=log_data.prompt,
        response=log_data.response,
        tokens_used=log_data.tokens_used,
        latency_ms=log_data.latency_ms,
        success=log_data.success,
        error_message=log_data.error_message,
        field_id=log_data.field_id,
        field_name=log_data.field_name,
        extra_data=json.dumps(log_data.extra_data) if log_data.extra_data else None,
        created_at=datetime.utcnow(),
    )

    db.add(log)
    db.commit()

    return {
        "success": True,
        "log_id": log.id,
        "message": "فعالیت ثبت شد"
    }


@router.get("/{project_id}/journal")
async def get_activity_journal(
    project_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    activity_type: Optional[str] = None,
    model_id: Optional[str] = None,
    success: Optional[bool] = None,
    sort_by: str = Query("created_at", pattern="^(created_at|tokens_used|latency_ms)$"),
    sort_order: str = Query("desc", pattern="^(asc|desc)$"),
    include_roadmap: bool = Query(False, description="اگر True باشد، اطلاعات نقشه راه هم برمی‌گردد"),  # 🔴 نقشه راه
    db: Session = Depends(get_db)
):
    """
    دریافت ژورنال فعالیت‌های پروژه با فیلتر و صفحه‌بندی

    🔴 نقشه راه فقط در تب ژورنال نمایش داده می‌شود.
    با پارامتر include_roadmap=true اطلاعات نقشه راه هم برمی‌گردد.
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    query = db.query(ActivityLog).filter(ActivityLog.project_id == project_id)

    # فیلترها
    if activity_type:
        query = query.filter(ActivityLog.activity_type == activity_type)
    if model_id:
        query = query.filter(ActivityLog.model_id == model_id)
    if success is not None:
        query = query.filter(ActivityLog.success == success)

    # تعداد کل
    total = query.count()

    # مرتب‌سازی
    sort_column = getattr(ActivityLog, sort_by)
    if sort_order == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(asc(sort_column))

    # صفحه‌بندی
    offset = (page - 1) * page_size
    logs = query.offset(offset).limit(page_size).all()

    result = {
        "success": True,
        "journal": [log_to_response(log) for log in logs],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
        }
    }

    # 🔴 اگر نقشه راه درخواست شده، آن را هم اضافه کن
    if include_roadmap and project.roadmap_content:
        roadmap_items = []
        lines = project.roadmap_content.split("\n")
        for line in lines:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            is_completed = "[x]" in line.lower() or "✅" in line
            if "[ ]" in line or "[x]" in line.lower() or line.startswith("-"):
                clean_text = line
                for prefix in ["- [ ]", "- [x]", "- [X]", "-", "[ ]", "[x]", "[X]"]:
                    clean_text = clean_text.replace(prefix, "")
                clean_text = clean_text.replace("✅", "").strip()
                if clean_text:
                    roadmap_items.append({"text": clean_text, "completed": is_completed})

        completed_count = sum(1 for i in roadmap_items if i["completed"])
        result["roadmap"] = {
            "items": roadmap_items,
            "total": len(roadmap_items),
            "completed": completed_count,
            "progress": round((completed_count / len(roadmap_items) * 100), 1) if roadmap_items else 0,
            "note": "نقشه راه فقط در تب ژورنال نمایش داده می‌شود"
        }

    return result


@router.get("/{project_id}/journal/roadmap")
async def get_roadmap_for_journal(
    project_id: str,
    db: Session = Depends(get_db)
):
    """
    🔴 دریافت نقشه راه برای نمایش در تب ژورنال
    نقشه راه فقط در تب ژورنال نمایش داده می‌شود (نه تب سلامت)
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    roadmap_content = project.roadmap_content or ""

    # پارس نقشه راه به آیتم‌های مجزا
    items = []
    lines = roadmap_content.split("\n")
    current_category = "general"

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # شناسایی دسته‌بندی
        if line.startswith("###") or line.startswith("##"):
            current_category = line.lstrip("#").strip()
            continue

        # شناسایی آیتم‌های چک‌لیست
        is_completed = "[x]" in line.lower() or "✅" in line
        is_checkbox = "[ ]" in line or "[x]" in line.lower()

        if is_checkbox or line.startswith("-") or line.startswith("*"):
            # پاکسازی متن
            clean_text = line
            for prefix in ["- [ ]", "- [x]", "- [X]", "-", "*", "[ ]", "[x]", "[X]"]:
                clean_text = clean_text.replace(prefix, "")
            clean_text = clean_text.replace("✅", "").strip()

            if clean_text:
                items.append({
                    "text": clean_text,
                    "completed": is_completed,
                    "category": current_category,
                })

    # تعداد آیتم‌ها
    completed_count = sum(1 for i in items if i["completed"])
    total_count = len(items)
    progress = round((completed_count / total_count * 100), 1) if total_count > 0 else 0

    # دریافت آخرین به‌روزرسانی نقشه راه از ژورنال
    last_roadmap_update = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.activity_type.in_(["engineering_report", "auto_setup_complete"])
    ).order_by(desc(ActivityLog.created_at)).first()

    return {
        "success": True,
        "project_id": project_id,
        "roadmap": {
            "raw_content": roadmap_content,
            "items": items,
            "total_items": total_count,
            "completed_items": completed_count,
            "progress_percent": progress,
            "last_updated": last_roadmap_update.created_at.isoformat() if last_roadmap_update else None,
            "last_updated_by": last_roadmap_update.model_id if last_roadmap_update else None,
        },
        "categories": list(set(i["category"] for i in items)),
        "note": "نقشه راه فقط در تب ژورنال نمایش داده می‌شود"
    }


@router.get("/{project_id}/journal/{log_id}/operations")
async def get_detailed_operations(
    project_id: str,
    log_id: str,
    db: Session = Depends(get_db)
):
    """
    🆕 دریافت عملیات جزئی یک ردیف ژورنال
    با کلیک روی هر ردیف، لیست عملیات جزئی آن نمایش داده می‌شود
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    # دریافت ردیف اصلی
    parent_log = db.query(ActivityLog).filter(ActivityLog.id == log_id).first()
    if not parent_log:
        raise HTTPException(status_code=404, detail="ردیف ژورنال یافت نشد")

    # دریافت عملیات جزئی
    operations = db.query(DetailedOperation).filter(
        DetailedOperation.parent_log_id == log_id
    ).order_by(DetailedOperation.sequence_number).all()

    return {
        "success": True,
        "log_id": log_id,
        "log_summary": {
            "activity_type": parent_log.activity_type,
            "model_id": parent_log.model_id,
            "field_name": parent_log.field_name,
            "created_at": parent_log.created_at.isoformat() if parent_log.created_at else None,
        },
        "operations": [detailed_op_to_response(op) for op in operations],
        "total_operations": len(operations)
    }


@router.get("/{project_id}/journal/operation/{operation_id}")
async def get_operation_details(
    project_id: str,
    operation_id: str,
    db: Session = Depends(get_db)
):
    """
    🆕 دریافت جزئیات کامل یک عملیات با کلیک
    شامل مقادیر قبل و بعد، جزئیات کامل و غیره
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    operation = db.query(DetailedOperation).filter(
        DetailedOperation.id == operation_id,
        DetailedOperation.project_id == project_id
    ).first()

    if not operation:
        raise HTTPException(status_code=404, detail="عملیات یافت نشد")

    # جزئیات کامل بدون truncate
    details = {}
    try:
        if operation.details:
            details = json.loads(operation.details)
    except:
        pass

    return {
        "success": True,
        "operation": {
            "id": operation.id,
            "project_id": operation.project_id,
            "parent_log_id": operation.parent_log_id,
            "sequence_number": operation.sequence_number,
            "operation_type": operation.operation_type,
            "summary": operation.summary,
            "details": details,
            "before_value": operation.before_value,  # کامل بدون truncate
            "after_value": operation.after_value,  # کامل بدون truncate
            "target_type": operation.target_type,
            "target_id": operation.target_id,
            "target_name": operation.target_name,
            "status": operation.status,
            "created_at": operation.created_at.isoformat() if operation.created_at else None,
        }
    }


@router.get("/{project_id}/journal/stats")
async def get_journal_stats(
    project_id: str,
    days: int = Query(7, ge=1, le=365),
    db: Session = Depends(get_db)
):
    """آمار ژورنال پروژه"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    since = datetime.utcnow() - timedelta(days=days)

    logs = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since
    ).all()

    # آمار
    total_activities = len(logs)
    total_tokens = sum(log.tokens_used for log in logs)
    total_latency = sum(log.latency_ms for log in logs)
    success_count = sum(1 for log in logs if log.success)

    # گروه‌بندی بر اساس مدل
    by_model = {}
    for log in logs:
        if log.model_id not in by_model:
            by_model[log.model_id] = {"count": 0, "tokens": 0}
        by_model[log.model_id]["count"] += 1
        by_model[log.model_id]["tokens"] += log.tokens_used

    # گروه‌بندی بر اساس نوع فعالیت
    by_type = {}
    for log in logs:
        if log.activity_type not in by_type:
            by_type[log.activity_type] = 0
        by_type[log.activity_type] += 1

    return {
        "success": True,
        "stats": {
            "period_days": days,
            "total_activities": total_activities,
            "total_tokens": total_tokens,
            "avg_latency_ms": total_latency // total_activities if total_activities > 0 else 0,
            "success_rate": round(success_count / total_activities * 100, 1) if total_activities > 0 else 0,
            "by_model": by_model,
            "by_type": by_type,
        }
    }


@router.get("/{project_id}/journal/{log_id}")
async def get_activity_detail(
    project_id: str,
    log_id: str,
    db: Session = Depends(get_db)
):
    """دریافت جزئیات کامل یک فعالیت"""
    log = db.query(ActivityLog).filter(
        ActivityLog.id == log_id,
        ActivityLog.project_id == project_id
    ).first()

    if not log:
        raise HTTPException(status_code=404, detail="فعالیت یافت نشد")

    return {
        "success": True,
        "activity": log_to_full_response(log)
    }


# ===================== گزارشات =====================

@router.get("/{project_id}/reports")
async def get_project_reports(
    project_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db)
):
    """دریافت لیست گزارشات پروژه"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    query = db.query(Report).filter(Report.project_id == project_id).order_by(desc(Report.created_at))

    total = query.count()
    offset = (page - 1) * page_size
    reports = query.offset(offset).limit(page_size).all()

    return {
        "success": True,
        "reports": [{
            "id": r.id,
            "report_type": r.report_type,
            "title": r.title,
            "summary": r.summary[:200] + "..." if r.summary and len(r.summary) > 200 else r.summary,
            "total_activities": r.total_activities,
            "total_tokens": r.total_tokens,
            "period_start": r.period_start.isoformat() if r.period_start else None,
            "period_end": r.period_end.isoformat() if r.period_end else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        } for r in reports],
        "pagination": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": (total + page_size - 1) // page_size,
        }
    }


@router.post("/{project_id}/reports/generate")
async def generate_report(
    project_id: str,
    days: int = Query(7, ge=1, le=30),
    model_id: str = Query("openai"),
    db: Session = Depends(get_db)
):
    """تولید گزارش ساده از فعالیت‌های اخیر"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    since = datetime.utcnow() - timedelta(days=days)

    logs = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since
    ).order_by(ActivityLog.created_at).all()

    if not logs:
        return {
            "success": False,
            "message": "فعالیتی در این بازه یافت نشد"
        }

    # آمار
    total_tokens = sum(log.tokens_used for log in logs)
    models_used = list(set(log.model_id for log in logs))

    # ساخت خلاصه فعالیت‌ها
    activities_summary = []
    for log in logs[:50]:  # حداکثر 50 فعالیت
        activities_summary.append({
            "type": log.activity_type,
            "model": log.model_id,
            "field": log.field_name,
            "prompt_preview": log.prompt[:100] if log.prompt else None,
            "success": log.success,
            "date": log.created_at.strftime("%Y-%m-%d %H:%M"),
        })

    # ایجاد گزارش
    report = Report(
        id=f"report_{uuid.uuid4().hex[:12]}",
        project_id=project_id,
        report_type="custom",
        title=f"گزارش {days} روز اخیر - {project.name}",
        content=json.dumps(activities_summary, ensure_ascii=False, indent=2),
        summary=f"تعداد {len(logs)} فعالیت با مصرف {total_tokens} توکن",
        total_activities=len(logs),
        total_tokens=total_tokens,
        models_used=json.dumps(models_used),
        period_start=since,
        period_end=datetime.utcnow(),
        created_at=datetime.utcnow(),
        generated_by=model_id,
    )

    db.add(report)
    db.commit()

    return {
        "success": True,
        "report_id": report.id,
        "message": f"گزارش با {len(logs)} فعالیت ایجاد شد"
    }


@router.post("/{project_id}/reports/generate-engineering")
async def generate_engineering_report(
    project_id: str,
    days: int = Query(7, ge=1, le=30),
    model_id: str = Query("claude"),
    auto_create_fields: bool = Query(True),
    validate_health_issues: bool = Query(True),  # 🆕 اعتبارسنجی نتایج health analysis
    db: Session = Depends(get_db)
):
    """
    تولید گزارش مهندسی جامع با:
    - تحلیل کامل ساختار پروژه
    - 🆕 اعتبارسنجی نتایج آخرین health analysis توسط مدل‌های منتخب
    - شناسایی باگ‌ها و مشکلات فنی
    - پیشنهادات بهبود
    - نقشه راه توسعه
    - تولید خودکار فیلدها با مارکر اعتبارسنجی
    - 🆕 آرشیو مسائل رد شده برای مشاهده
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.ai_base import Message
    from ...models.project import ProjectFile
    import logging
    logger = logging.getLogger(__name__)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    # دریافت فعالیت‌های اخیر
    since = datetime.utcnow() - timedelta(days=days)
    logs = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since
    ).order_by(ActivityLog.created_at).all()

    # دریافت تنظیمات محدودیت‌ها
    from .settings import get_ai_limits_sync
    ai_limits = get_ai_limits_sync(db)
    limits_enabled = ai_limits.get("limits_enabled", False)

    # مقادیر محدودیت (0 = نامحدود)
    max_files = ai_limits.get("max_files_for_report", 0) if limits_enabled else 0
    max_code_samples = ai_limits.get("max_code_samples_for_report", 0) if limits_enabled else 0
    max_chars_per_sample = ai_limits.get("max_chars_per_code_sample", 0) if limits_enabled else 0

    # دریافت فایل‌های پروژه
    files = db.query(ProjectFile).filter(ProjectFile.project_id == project_id).all()
    files_summary = []
    code_samples = []
    total_code_chars = 0

    # 🔴 بررسی عمیق GitHub: اگر پروژه از GitHub است، فایل‌های جدید را دریافت کن
    github_deep_inspection = False
    github_inspection_result = None
    if project.project_type == "github_import":
        try:
            extra_data = json.loads(project.extra_data) if project.extra_data else {}
            owner = extra_data.get("owner")
            repo = extra_data.get("repo")
            if owner and repo:
                logger.info(f"🔍 Deep GitHub inspection for {owner}/{repo}")

                # دریافت محتوای جدید از GitHub با API
                import httpx
                try:
                    async with httpx.AsyncClient() as client:
                        # اول سینک کن تا فایل‌های جدید دریافت شوند
                        sync_response = await client.post(
                            f"http://localhost:8000/api/github/imported/{project_id}/refresh",
                            timeout=30.0
                        )
                        if sync_response.status_code == 200:
                            sync_data = sync_response.json()
                            github_deep_inspection = True
                            github_inspection_result = {
                                "synced": True,
                                "files_added": sync_data.get("files_added", 0),
                                "files_updated": sync_data.get("files_updated", 0),
                                "files_removed": sync_data.get("files_removed", 0),
                            }
                            logger.info(f"✅ GitHub sync completed: {github_inspection_result}")

                            # دوباره فایل‌ها را بخوان چون ممکنه به‌روز شده باشند
                            db.expire_all()
                            files = db.query(ProjectFile).filter(ProjectFile.project_id == project_id).all()
                except Exception as e:
                    logger.warning(f"GitHub deep inspection failed: {e}")
                    github_inspection_result = {"synced": False, "error": str(e)}
        except Exception as e:
            logger.warning(f"Error checking GitHub info: {e}")

    # اولویت‌بندی فایل‌ها - فایل‌های مهم‌تر اول
    priority_files = ['auth', 'login', 'user', 'route', 'api', 'main', 'app', 'index', 'config', 'setting']
    code_extensions = ['py', 'ts', 'tsx', 'js', 'jsx', 'vue', 'svelte', 'java', 'go', 'rs', 'rb', 'php']

    # مرتب‌سازی فایل‌ها بر اساس اهمیت
    def file_priority(f):
        name = f.file_path.lower()
        score = 0
        for pf in priority_files:
            if pf in name:
                score += 10
        if f.file_type in code_extensions:
            score += 5
        return -score  # منفی برای مرتب‌سازی نزولی

    sorted_files = sorted(files, key=file_priority)

    files_added = 0
    for f in sorted_files:
        # چک محدودیت تعداد فایل (0 = نامحدود)
        if max_files > 0 and files_added >= max_files:
            break

        files_summary.append({
            "path": f.file_path,
            "type": f.file_type,
            "size": len(f.content) if f.content else 0
        })
        files_added += 1

        # نمونه کد از فایل‌های کد
        if f.content and f.file_type in code_extensions:
            # چک محدودیت تعداد نمونه کد (0 = نامحدود)
            if max_code_samples > 0 and len(code_samples) >= max_code_samples:
                continue

            # محتوای فایل (0 = بدون محدودیت)
            content = f.content if max_chars_per_sample == 0 else f.content[:max_chars_per_sample]
            code_samples.append({
                "path": f.file_path,
                "content": content
            })
            total_code_chars += len(content)

    # دریافت فیلدهای فعلی
    existing_fields = []
    try:
        if project.dynamic_fields:
            existing_fields = json.loads(project.dynamic_fields)
    except:
        pass

    # ====================================
    # 🆕 دریافت و آماده‌سازی نتایج health analysis برای اعتبارسنجی
    # ====================================
    health_analysis_issues = []
    health_analysis_summary = ""
    partial_results = {}

    # 🔴 DEBUG: Log raw data before extraction
    logger.info(f"=" * 60)
    logger.info(f"🔴 DEBUG: Starting health issues extraction for project {project_id}")
    logger.info(f"🔴 DEBUG: validate_health_issues = {validate_health_issues}")
    logger.info(f"🔴 DEBUG: project.issues_found type: {type(project.issues_found)}")
    logger.info(f"🔴 DEBUG: project.issues_found length: {len(project.issues_found or '')}")
    if project.issues_found:
        logger.info(f"🔴 DEBUG: project.issues_found preview: {project.issues_found[:500]}...")
    else:
        # 🔴 CRITICAL: Try to refresh from database!
        logger.error(f"🔴🔴🔴 CRITICAL: project.issues_found is None/empty! Trying to refresh from DB...")
        db.refresh(project)
        logger.info(f"🔴 DEBUG: After refresh: project.issues_found length: {len(project.issues_found or '')}")
        if project.issues_found:
            logger.info(f"🔴 DEBUG: After refresh preview: {project.issues_found[:500]}...")
    logger.info(f"=" * 60)

    def normalize_issue(issue, file_path=None, source_models=None):
        """نرمال‌سازی فرمت ایراد از منابع مختلف"""
        return {
            "file": issue.get("file") or file_path or "unknown",
            "type": issue.get("type") or issue.get("category") or "code_quality",
            "severity": issue.get("severity") or "medium",
            "message": issue.get("message") or issue.get("description") or issue.get("problem") or str(issue),
            "line": issue.get("line"),
            "source_models": source_models or issue.get("source_models") or ["unknown"],
        }

    # 1. دریافت از file_health_map (نتایج تحلیل هر فایل)
    if project.file_health_map:
        try:
            file_map = json.loads(project.file_health_map)
            for file_path, file_data in file_map.items():
                if isinstance(file_data, dict):
                    # استخراج issues از داخل file_data
                    file_issues = file_data.get("issues", [])
                    if isinstance(file_issues, list):
                        for issue in file_issues:
                            if isinstance(issue, dict):
                                health_analysis_issues.append(normalize_issue(
                                    issue,
                                    file_path=file_path,
                                    source_models=file_data.get("analyzed_by", ["file_analysis"])
                                ))
        except Exception as e:
            logger.warning(f"Error parsing file_health_map: {e}")

    logger.info(f"🔴 DEBUG: After file_health_map extraction: {len(health_analysis_issues)} issues")

    # 2. دریافت نتایج از analysis_progress (partial_results)
    if project.analysis_progress:
        try:
            progress_data = json.loads(project.analysis_progress)
            partial_results = progress_data.get("partial_results", {})

            # استخراج issues از micro analysis
            if "micro" in partial_results:
                micro_results = partial_results["micro"]
                if isinstance(micro_results, dict):
                    for file_path, file_data in micro_results.items():
                        if isinstance(file_data, dict):
                            issues = file_data.get("issues", [])
                            if isinstance(issues, list):
                                for issue in issues:
                                    if isinstance(issue, dict):
                                        health_analysis_issues.append(normalize_issue(
                                            issue,
                                            file_path=file_path,
                                            source_models=file_data.get("analyzed_by", [])
                                        ))

            # استخراج issues از structural analysis
            if "structural" in partial_results:
                structural = partial_results["structural"]
                if isinstance(structural, dict):
                    for issue in structural.get("issues", []):
                        if isinstance(issue, dict):
                            health_analysis_issues.append(normalize_issue(
                                issue,
                                file_path=issue.get("file", "structural"),
                                source_models=["structural_analysis"]
                            ))
        except Exception as e:
            logger.warning(f"Error parsing analysis_progress: {e}")

    logger.info(f"🔴 DEBUG: After analysis_progress extraction: {len(health_analysis_issues)} issues")

    # 3. دریافت از issues_found (ذخیره شده از قبل) - این منبع اصلی است!
    logger.info(f"🔴 DEBUG: Checking project.issues_found...")
    if project.issues_found:
        try:
            stored_issues = json.loads(project.issues_found)
            logger.info(f"🔴 DEBUG: Parsed issues_found: type={type(stored_issues)}, count={len(stored_issues) if isinstance(stored_issues, list) else 'N/A'}")
            if isinstance(stored_issues, list):
                # ادغام با issues موجود (بدون تکرار)
                existing_keys = {f"{i.get('file', '')}:{i.get('line', '')}:{i.get('message', '')[:50]}" for i in health_analysis_issues}
                added_count = 0
                for issue in stored_issues:
                    if isinstance(issue, dict):
                        normalized = normalize_issue(issue)
                        key = f"{normalized['file']}:{normalized.get('line', '')}:{normalized['message'][:50]}"
                        if key not in existing_keys:
                            health_analysis_issues.append(normalized)
                            existing_keys.add(key)
                            added_count += 1
                logger.info(f"🔴 DEBUG: Added {added_count} issues from issues_found (after dedup)")
            else:
                logger.warning(f"🔴 DEBUG: issues_found is not a list! Content: {str(stored_issues)[:200]}")
        except Exception as e:
            logger.warning(f"Error parsing issues_found: {e}")
    else:
        logger.warning(f"🔴 DEBUG: project.issues_found is EMPTY!")

    logger.info(f"🔴 DEBUG: After issues_found extraction: {len(health_analysis_issues)} issues")

    # 4. دریافت از health_scores (ممکن است issues داخلش باشد)
    if project.health_scores:
        try:
            scores = json.loads(project.health_scores)
            if isinstance(scores, dict) and "issues" in scores:
                for issue in scores.get("issues", []):
                    if isinstance(issue, dict):
                        health_analysis_issues.append(normalize_issue(issue))
        except:
            pass

    # Log detailed extraction stats
    logger.info(f"🔍 Health issues extraction from project {project_id}:")
    logger.info(f"   - project.file_health_map: {len(project.file_health_map or '')} chars")
    logger.info(f"   - project.analysis_progress: {len(project.analysis_progress or '')} chars")
    logger.info(f"   - project.issues_found: {len(project.issues_found or '')} chars")
    logger.info(f"   - project.health_scores: {len(project.health_scores or '')} chars")
    logger.info(f"📊 Total health issues found for validation: {len(health_analysis_issues)}")

    # Group issues by severity for logging
    if health_analysis_issues:
        severity_counts = {}
        for issue in health_analysis_issues:
            sev = issue.get("severity", "unknown")
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
        logger.info(f"📊 Issues by severity: {severity_counts}")
    else:
        logger.warning(f"⚠️ No health issues found from any source. Attempting recovery...")
        # 🔴 FAILSAFE: Try to force-load issues from issues_found
        if project.issues_found:
            try:
                stored = json.loads(project.issues_found)
                if isinstance(stored, list) and len(stored) > 0:
                    logger.info(f"🔴 FAILSAFE: Found {len(stored)} issues in issues_found, force-loading...")
                    for issue in stored:
                        if isinstance(issue, dict):
                            health_analysis_issues.append(normalize_issue(issue))
                    logger.info(f"🔴 FAILSAFE: Loaded {len(health_analysis_issues)} issues from issues_found")
            except Exception as e:
                logger.error(f"🔴 FAILSAFE failed: {e}")

        # Additional debug for empty issues
        if not health_analysis_issues:
            logger.error(f"🔴🔴🔴 CRITICAL: Still no issues after failsafe!")
            if project.file_health_map:
                try:
                    fhm = json.loads(project.file_health_map)
                    total_issues_in_map = sum(len(fd.get("issues", [])) for fd in fhm.values() if isinstance(fd, dict))
                    total_issues_count = sum(fd.get("issues_count", 0) for fd in fhm.values() if isinstance(fd, dict))
                    logger.info(f"   - file_health_map has {len(fhm)} files, {total_issues_in_map} issues embedded, {total_issues_count} issues_count")
                except:
                    pass

    # ساخت خلاصه health analysis برای prompt
    if health_analysis_issues:
        health_analysis_summary = f"""

=== 🔍 نتایج آخرین health analysis ({len(health_analysis_issues)} ایراد شناسایی شده) ===

⚠️ بسیار مهم: این ایرادات توسط تحلیل سلامت شناسایی شده‌اند و باید توسط شما اعتبارسنجی شوند!
- هر ایراد را با کد واقعی مقایسه کن
- اگر ایراد واقعاً وجود دارد، در validated_issues قرار بده (با create_field=true)
- اگر ایراد اشتباه است، در rejected_issues با دلیل رد شدن قرار بده

لیست ایرادات برای بررسی (تمام {len(health_analysis_issues)} ایراد):
{json.dumps(health_analysis_issues, ensure_ascii=False, indent=2)}
"""
    else:
        health_analysis_summary = """

=== 🔍 نتایج تحلیل سلامت ===
هیچ ایرادی از تحلیل سلامت قبلی یافت نشد.
در صورتی که مشکلاتی در کد مشاهده می‌کنید، آنها را در بخش bugs_and_issues گزارش کنید.
"""

    # خلاصه فعالیت‌ها
    activities_summary = []
    for log in logs[:30]:
        activities_summary.append({
            "type": log.activity_type,
            "model": log.model_id,
            "field": log.field_name,
            "success": log.success,
            "error": log.error_message if not log.success else None,
            "date": log.created_at.strftime("%Y-%m-%d %H:%M"),
        })

    # ====================================
    # 🆕 استخراج ردیف‌های ژورنال برای بررسی
    # - فقط ردیف‌هایی که از آخرین گزارش جدیدتر هستند
    # - شامل محتوای response برای تحلیل نتایج بررسی‌ها
    # ====================================
    journal_entries_for_review = []
    last_report_date = None

    # یافتن آخرین گزارش مهندسی
    last_eng_report = db.query(Report).filter(
        Report.project_id == project_id,
        Report.report_type == "engineering"
    ).order_by(Report.created_at.desc()).first()

    if last_eng_report:
        last_report_date = last_eng_report.created_at
        logger.info(f"📋 Last engineering report: {last_report_date}")

    # فیلتر ژورنال‌های جدید
    journal_filter = ActivityLog.project_id == project_id
    if last_report_date:
        journal_filter = journal_filter & (ActivityLog.created_at > last_report_date)

    new_journal_entries = db.query(ActivityLog).filter(
        journal_filter
    ).order_by(ActivityLog.created_at.desc()).limit(50).all()

    logger.info(f"📋 Found {len(new_journal_entries)} new journal entries since last report")

    # استخراج محتوای مهم از ژورنال‌ها
    for entry in new_journal_entries:
        # فقط ورودی‌هایی که response دارند و مرتبط با بررسی/تحلیل هستند
        if entry.response and len(entry.response) > 50:
            entry_data = {
                "id": entry.id,
                "type": entry.activity_type,
                "model": entry.model_id,
                "field_name": entry.field_name,
                "prompt_summary": entry.prompt[:200] if entry.prompt else "",
                "response_summary": entry.response[:1500],  # 🆕 شامل محتوای response
                "success": entry.success,
                "date": entry.created_at.strftime("%Y-%m-%d %H:%M"),
                "tokens_used": entry.tokens_used,
            }

            # استخراج نتایج بررسی از response (اگر بررسی/تحلیل بود)
            if entry.activity_type in ["trigger", "analysis", "engineering_report"]:
                entry_data["is_review"] = True
                # جستجوی کلیدواژه‌های نتایج بررسی
                review_keywords = ["مشکل", "ایراد", "خطا", "پیشنهاد", "اصلاح", "بهبود", "error", "issue", "bug", "fix", "improve"]
                entry_data["has_findings"] = any(kw in (entry.response or "").lower() for kw in review_keywords)
            else:
                entry_data["is_review"] = False
                entry_data["has_findings"] = False

            journal_entries_for_review.append(entry_data)

    # گروه‌بندی ژورنال‌ها براساس نوع
    journal_reviews = [e for e in journal_entries_for_review if e.get("is_review") and e.get("has_findings")]
    journal_other = [e for e in journal_entries_for_review if not e.get("is_review")]

    logger.info(f"📋 Journal entries with review findings: {len(journal_reviews)}")

    # 🔴 تلاش برای دریافت پرامپت از دیتابیس
    db_system_prompt = PromptHelper.get_prompt(
        db=db,
        category="engineering_report",
        prompt_id="eng_system_prompt",  # ID مطابق با seed data
        variables={}
    )

    if db_system_prompt:
        logger.info("📝 Using DB prompt for engineering report system prompt")
        system_prompt = db_system_prompt
    else:
        logger.debug("📝 Using hardcoded engineering report system prompt")
        # 🔄 Fallback به پرامپت hardcoded
        # ساخت prompt برای AI
        system_prompt = """تو یک مهندس ارشد نرم‌افزار هستی که باید یک گزارش مهندسی جامع و حرفه‌ای تولید کنی.

🔴🔴🔴 بسیار مهم - اعتبارسنجی health analysis 🔴🔴🔴
اگر در بخش ورودی "نتایج آخرین health analysis" وجود دارد، باید:
1. هر ایراد را یک به یک بررسی کنی
2. ایرادات تایید شده را در validated_issues قرار بدی
3. ایرادات رد شده را در rejected_issues قرار بدی
4. بخش health_analysis_validation را حتماً در JSON خروجی قرار بدی

گزارش باید شامل بخش‌های زیر باشد (حتماً از این ساختار JSON استفاده کن):

```json
{
    "executive_summary": "خلاصه مدیریتی 2-3 پاراگراف از وضعیت پروژه",

    "project_health": {
        "score": 75,
        "status": "متوسط/خوب/عالی/نیاز به توجه",
        "key_metrics": {
            "code_quality": 70,
            "documentation": 50,
            "test_coverage": 30,
            "architecture": 80
        }
    },

    "technical_analysis": {
        "strengths": ["نقطه قوت 1", "نقطه قوت 2"],
        "weaknesses": ["نقطه ضعف 1", "نقطه ضعف 2"],
        "architecture_review": "تحلیل معماری",
        "code_quality_issues": [
            {"file": "path/to/file", "issue": "توضیح مشکل", "severity": "high/medium/low"}
        ]
    },

    "bugs_and_issues": [
        {"title": "عنوان باگ", "description": "توضیحات", "severity": "critical/high/medium/low", "file": "path/to/file", "suggested_fix": "راه حل پیشنهادی"}
    ],

    "security_review": {
        "vulnerabilities": [{"type": "نوع آسیب‌پذیری", "location": "محل", "risk": "high/medium/low"}],
        "recommendations": ["پیشنهاد امنیتی 1", "پیشنهاد امنیتی 2"]
    },

    "performance_analysis": {
        "bottlenecks": ["گلوگاه 1", "گلوگاه 2"],
        "optimization_suggestions": ["پیشنهاد بهینه‌سازی 1"]
    },

    "recommendations": [
        {"priority": "high/medium/low", "title": "عنوان", "description": "توضیحات", "effort": "کم/متوسط/زیاد"}
    ],

    "field_management": {
        "fields_to_archive": ["id_فیلدهایی که انجام شده یا دیگر نیاز نیست"],
        "fields_to_approve": ["id_فیلدهای pending که تایید می‌شوند برای اجرا"],
        "fields_to_reject": ["id_فیلدهای pending که رد و بایگانی می‌شوند"],
        "fields_to_merge": [
            {"source_ids": ["id1", "id2"], "merged_name": "نام جدید", "merged_value": "دستور ادغام‌شده"}
        ],
        "fields_to_update": [
            {"id": "id_فیلد", "new_value": "دستور جدید", "new_priority": 3}
        ]
    },

    "roadmap": {
        "immediate": [
            {"task": "کار فوری 1", "description": "توضیحات کامل برای تولید کد", "action_type": "github_commit", "target_path": "path/to/file.py", "priority": 1, "field_type": "temporary"}
        ],
        "short_term": [
            {"task": "کار کوتاه‌مدت 1", "description": "توضیحات", "action_type": "github_commit", "target_path": "path/to/file.py", "priority": 3, "field_type": "temporary"}
        ],
        "long_term": [
            {"task": "کار بلندمدت 1", "description": "توضیحات", "action_type": "display", "priority": 7, "field_type": "permanent"}
        ]
    },

    "roadmap_status_updates": [
        {"item": "نام آیتم نقشه راه", "completed": true, "reason": "فیلد X اجرا شده و مشکل حل شده"},
        {"item": "نام آیتم نقشه راه", "completed": false, "create_field": true, "field_details": {"name": "...", "value": "...", "target_path": "...", "action_type": "github_commit"}}
    ],

    "activity_analysis": {
        "success_rate": 75,
        "failed_tasks_analysis": "تحلیل تسک‌های ناموفق",
        "model_performance": {"claude": "خوب", "openai": "متوسط"}
    },

    "health_analysis_validation": {
        "total_reviewed": 15,
        "validated_issues": [
            {
                "original_issue": {"file": "path/file.py", "type": "security", "message": "..."},
                "validation_score": 95,
                "validation_note": "این ایراد تایید می‌شود چون...",
                "priority": "high",
                "create_field": true
            }
        ],
        "rejected_issues": [
            {
                "original_issue": {"file": "path/file.py", "type": "unused", "message": "..."},
                "rejection_reason": "این متغیر در خط 45 استفاده شده است",
                "validation_score": 20,
                "source_model_error": "مدل X اشتباهاً این را ایراد تشخیص داده"
            }
        ],
        "validation_summary": "خلاصه اعتبارسنجی: از 15 ایراد، 10 تایید و 5 رد شد"
    },

    "comprehensive_ideal_state": {
        "description": "توضیح کامل حالت ایده‌آل پروژه (3-5 پاراگراف)",
        "current_deficiencies": ["کمبود 1", "کمبود 2"],
        "unexecuted_tasks": ["تسک اجرا نشده 1", "تسک اجرا نشده 2"],
        "system_structure": {
            "overview": "ساختار کلی سیستم",
            "components": ["کامپوننت 1", "کامپوننت 2"],
            "wiring": "نحوه اتصال و سیم‌کشی بین بخش‌ها"
        },
        "roadmap_integration": "نحوه رسیدن از وضعیت فعلی به حالت ایده‌آل"
    }
}
```

⚠️ مهم:
- فقط JSON خروجی بده، بدون هیچ توضیح اضافی
- در roadmap، برای هر تسک که نیاز به تولید کد دارد، action_type را "github_commit" و target_path را مسیر فایل هدف قرار بده
- توضیحات هر تسک باید به قدری کامل باشد که AI بتواند مستقیماً کد تولید کند
- در field_management، فیلدهای موجود را بررسی کن و تصمیم بگیر کدام‌ها بایگانی، ادغام یا به‌روزرسانی شوند
- priority از 1 (بالاترین) تا 10 (پایین‌ترین): 1-2=فوری، 3-4=بالا، 5=عادی، 6-7=پایین، 8-10=خیلی پایین
- field_type: "permanent" برای دائمی/تکرارشونده، "temporary" برای یکبار مصرف

🔴 اعتبارسنجی health analysis (بسیار مهم):
- تمام ایرادات ارسال شده از health analysis را یک به یک بررسی کن
- با نگاه به کد واقعی، مشخص کن آیا هر ایراد معتبر است یا نه
- ایرادات معتبر را در validated_issues با create_field=true قرار بده (فیلد ساخته می‌شود)
- ایرادات نامعتبر را در rejected_issues با دلیل دقیق رد شدن قرار بده
- validation_score از 0-100: بالای 70 = معتبر، زیر 30 = رد شده

🔴🔴🔴 تأیید اصلاحات قبلی (بسیار مهم):
- فیلدهای بایگانی شده (archived/executed) را بررسی کن
- با نگاه به کد فعلی، تأیید کن که مشکلات واقعاً حل شده‌اند
- اگر مشکلی هنوز در کد وجود دارد، دوباره در validated_issues با یادداشت "مشکل هنوز حل نشده" قرار بده
- این شامل: خطاهای runtime، مشکلات authentication/login، ارتباط frontend-backend

🔴 شناسایی مشکلات واقعی runtime:
- به دنبال مشکلات احراز هویت (login, auth, session) بگرد
- ارتباط frontend با backend API را بررسی کن
- خطاهای CORS, 401, 403, 500 را شناسایی کن
- مسیرهای API که ممکن است مشکل داشته باشند را پیدا کن
- اگر endpoint تعریف نشده یا اشتباه است، گزارش بده

🟢 حالت ایده‌آل جامع (comprehensive_ideal_state):
- وضعیت ایده‌آل باید شامل: کمبودها، تسک‌های اجرا نشده، ساختار سیستم، سیم‌کشی و نقشه راه باشد
- این بخش برای راهنمایی توسعه‌دهنده بسیار مهم است

🔵🔵🔵 بررسی ردیف‌های ژورنال (بسیار مهم):
- ردیف‌های جدید ژورنال که محتوای بررسی/تحلیل دارند را بخوان
- نتایج بررسی هر ردیف را استخراج کن (مشکلات شناسایی شده، پیشنهادات)
- برای هر مشکل مهم شناسایی شده، یک فیلد عملیاتی ایجاد کن
- فیلدها باید action_type=github_commit یا github_multi_commit داشته باشند (نه display!)
- فیلدهای صرفاً نمایشی فایده‌ای ندارند - فیلدها باید کار انجام دهند

📝 ایجاد فیلدهای عملیاتی (بسیار مهم):
- هر فیلد باید action_type داشته باشد: github_commit (برای تغییر یک فایل) یا github_multi_commit (برای چند فایل)
- target_path حتماً باید مسیر فایل هدف باشد
- توضیحات فیلد باید آنقدر کامل باشد که AI بتواند مستقیماً کد تولید کند
- فیلدهای display فقط برای موارد مشاوره‌ای یا تحلیلی که نیاز به کد ندارند

ساختار خروجی journal_analysis (اگر ردیف‌های جدید وجود دارد):
```json
"journal_analysis": {
    "entries_reviewed": 10,
    "findings": [
        {
            "journal_entry_id": 123,
            "field_name": "نام فیلدی که این بررسی را تولید کرده",
            "finding_type": "bug|issue|suggestion|improvement",
            "summary": "خلاصه یافته",
            "severity": "critical|high|medium|low",
            "create_actionable_field": true,
            "suggested_field": {
                "name": "نام فیلد عملیاتی",
                "value": "دستور دقیق برای تولید کد و رفع مشکل",
                "action_type": "github_commit",
                "target_path": "path/to/file.py",
                "priority": 2
            }
        }
    ],
    "summary": "خلاصه بررسی ژورنال‌های جدید"
}
```"""

    # ساخت خلاصه فیلدهای اجرا نشده برای ideal state
    unexecuted_fields = [
        {"name": f.get("name"), "value": f.get("value", "")[:200], "priority": f.get("priority", 5)}
        for f in existing_fields
        if not f.get("archived") and f.get("field_type") == "temporary" and not f.get("executed")
    ]

    # 🆕 تهیه متن ژورنال برای پرامپت
    journal_section = ""
    if journal_reviews:
        journal_section = f"""

=== 🔵 ردیف‌های ژورنال جدید با یافته‌های مهم ({len(journal_reviews)} عدد) ===
این ردیف‌ها از آخرین گزارش مهندسی جدیدتر هستند و نتایج بررسی/تحلیل دارند.
برای هر یافته مهم، یک فیلد عملیاتی (با action_type=github_commit) ایجاد کن!

{json.dumps(journal_reviews[:15], ensure_ascii=False, indent=2)}
"""
    elif new_journal_entries:
        journal_section = f"""

=== 🔵 ردیف‌های ژورنال جدید ({len(new_journal_entries)} عدد) ===
{json.dumps([{"id": e.get("id"), "type": e.get("type"), "field_name": e.get("field_name"), "date": e.get("date")} for e in journal_entries_for_review[:20]], ensure_ascii=False, indent=2)}
"""

    user_prompt = f"""پروژه: {project.name}
توضیحات: {project.description or 'ندارد'}
نوع پروژه: {project.project_type or 'نامشخص'}

=== ساختار فایل‌ها ===
{json.dumps(files_summary, ensure_ascii=False, indent=2)}

=== کدهای پروژه ({len(code_samples)} فایل، {total_code_chars:,} کاراکتر) ===
{json.dumps(code_samples, ensure_ascii=False, indent=2)}
{health_analysis_summary if validate_health_issues else ''}
=== فعالیت‌های اخیر ({days} روز) ===
{json.dumps(activities_summary, ensure_ascii=False, indent=2)}
{journal_section}
=== 🔴 فیلدهای PENDING (نیاز به تایید مهندسی - بسیار مهم!) ===
{json.dumps([{"id": f.get("id"), "name": f.get("name"), "value": f.get("value", "")[:300], "action_type": f.get("action_type"), "target_path": f.get("target_path"), "source": f.get("source", "unknown"), "priority": f.get("priority", 5)} for f in existing_fields if not f.get("archived") and not f.get("engineering_approval")], ensure_ascii=False, indent=2)}

=== فیلدهای تایید شده (دارای تاییدیه مهندسی) ===
{json.dumps([{"id": f.get("id"), "name": f.get("name"), "action_type": f.get("action_type"), "executed": f.get("executed", False)} for f in existing_fields if not f.get("archived") and f.get("engineering_approval")], ensure_ascii=False, indent=2)}

=== فیلدهای بایگانی/اجرا شده ===
{json.dumps([{"id": f.get("id"), "name": f.get("name"), "target_path": f.get("target_path")} for f in existing_fields if f.get("archived") or f.get("executed")][:15], ensure_ascii=False, indent=2)}

=== حالت ایده‌آل فعلی ===
{project.ideal_state or 'تعریف نشده'}

=== 🗺️ نقشه راه (آیتم‌های ناقص باید به فیلد تبدیل شوند!) ===
{(project.roadmap_content or '')[:3000] if project.roadmap_content else 'تعریف نشده'}

لطفاً گزارش مهندسی جامع تولید کن.

🔴🔴🔴 وظایف اجباری - حتماً انجام بده:

1️⃣ **اعتبارسنجی فیلدهای PENDING:**
   - هر فیلد pending را بررسی کن
   - اگر لازم است: id را در fields_to_approve قرار بده
   - اگر غیرضروری/تکراری: id را در fields_to_reject قرار بده
   - اگر قابل ادغام با فیلد دیگر: در fields_to_merge قرار بده

2️⃣ **اعتبارسنجی health analysis:**
   - اگر بخش "نتایج health analysis" بالا وجود دارد، تمام ایرادات را بررسی کن
   - ایرادات تایید شده در validated_issues
   - ایرادات رد شده در rejected_issues

3️⃣ **بررسی نقشه راه:**
   - آیتم‌های انجام شده را با ✅ علامت بزن (در roadmap_status_updates با completed=true)
   - برای آیتم‌های انجام نشده، یک فیلد عملیاتی ایجاد کن (create_field=true)

4️⃣ **بررسی ژورنال:**
   - ردیف‌های جدید ژورنال را بررسی کن
   - برای یافته‌های مهم، فیلد عملیاتی با action_type=github_commit ایجاد کن

⚠️ یادآوری:
- فیلدهای display ایجاد نکن - فقط فیلدهای عملیاتی (github_commit/github_multi_commit)
- بدون تاییدیه مهندسی، هیچ فیلدی قابل اجرا نیست"""

    # فراخوانی AI
    ai_manager = get_ai_manager()

    # 🔴 بررسی فعال بودن مدل و یافتن جایگزین در صورت نیاز
    original_model_id = model_id
    used_fallback = False

    if not ai_manager.get_enabled_status(model_id):
        logger.warning(f"Model {model_id} is disabled, looking for fallback...")
        fallback_model = ai_manager.find_fallback_model(model_id, task_type="engineering_report")
        if fallback_model:
            logger.info(f"Using fallback model: {fallback_model} instead of {model_id}")
            model_id = fallback_model
            used_fallback = True
        else:
            # هیچ مدل فعالی نیست - سعی کن با لیست مدل‌های فعال
            available = ai_manager.get_available_models(task_type="engineering_report")
            if available:
                model_id = available[0].id
                used_fallback = True
                logger.info(f"No specific fallback, using first available: {model_id}")
            else:
                raise HTTPException(status_code=400, detail="هیچ مدل AI فعالی در سیستم وجود ندارد")

    # 🔴 تخمین توکن و کوتاه کردن prompt در صورت نیاز
    # تقریباً هر 4 کاراکتر = 1 توکن (برای متن فارسی/انگلیسی ترکیبی)
    estimated_tokens = (len(system_prompt) + len(user_prompt)) // 3

    # دریافت context_window مدل
    from ...core.models_registry import get_model
    model_info = get_model(model_id)
    max_context = model_info.context_window if model_info else 200000

    # حداکثر 80% از context window برای prompt (20% برای output)
    max_prompt_tokens = int(max_context * 0.80)

    logger.info(f"Estimated tokens: {estimated_tokens}, Max allowed: {max_prompt_tokens}")

    # اگر بیش از حد بود، prompt را کوتاه کن
    if estimated_tokens > max_prompt_tokens:
        logger.warning(f"Prompt too long ({estimated_tokens} tokens), truncating...")

        # کوتاه کردن user_prompt (system_prompt را دست نمی‌زنیم)
        excess_chars = (estimated_tokens - max_prompt_tokens) * 3

        # اول سعی کن code_samples را کوتاه کنی (معمولاً بیشترین حجم)
        if len(user_prompt) > excess_chars:
            user_prompt = user_prompt[:len(user_prompt) - excess_chars - 500]
            user_prompt += "\n\n[... محتوا به دلیل محدودیت توکن کوتاه شد ...]"
            logger.info(f"Truncated user_prompt to {len(user_prompt)} chars")

        # محاسبه مجدد
        estimated_tokens = (len(system_prompt) + len(user_prompt)) // 3
        logger.info(f"After truncation: {estimated_tokens} tokens")

    messages = [
        Message(role="system", content=system_prompt),
        Message(role="user", content=user_prompt),
    ]

    # 🔴 DEBUG: Final summary before AI call
    logger.info(f"=" * 60)
    logger.info(f"🔴 DEBUG: FINAL SUMMARY before AI call:")
    logger.info(f"   - model_id: {model_id} (original: {original_model_id}, fallback: {used_fallback})")
    logger.info(f"   - validate_health_issues: {validate_health_issues}")
    logger.info(f"   - health_analysis_issues count: {len(health_analysis_issues)}")
    logger.info(f"   - health_analysis_summary length: {len(health_analysis_summary)}")
    logger.info(f"   - health_analysis_summary included in prompt: {'yes' if (validate_health_issues and len(health_analysis_issues) > 0) else 'NO!'}")
    logger.info(f"   - user_prompt length: {len(user_prompt)}")
    logger.info(f"   - system_prompt length: {len(system_prompt)}")
    logger.info(f"   - estimated_tokens: {estimated_tokens}")
    if len(health_analysis_issues) == 0:
        logger.error(f"🔴🔴🔴 CRITICAL: No health issues to validate! Check extraction logic above.")
    logger.info(f"=" * 60)

    try:
        response = await ai_manager.generate(
            model_id=model_id,
            messages=messages,
            max_tokens=8192,
            temperature=0.3,
            task_type="engineering_report",
            allow_fallback=True,
        )

        # پارس JSON از پاسخ
        report_content = response.content
        report_data = None

        # استخراج JSON از پاسخ - بهبود یافته برای مدیریت markdown code blocks
        import re

        def try_fix_json(json_str):
            """تلاش برای تصحیح خطاهای رایج JSON"""
            fixes = [
                # حذف trailing commas قبل از } یا ]
                (r',(\s*[}\]])', r'\1'),
                # حذف trailing comma در آخر
                (r',\s*$', ''),
                # تصحیح newlines در strings
                (r'(?<!\\)\n(?=[^"]*"[^"]*$)', r'\\n'),
            ]

            fixed = json_str
            for pattern, replacement in fixes:
                fixed = re.sub(pattern, replacement, fixed)
            return fixed

        def attempt_json_parse(json_str, source_name):
            """تلاش برای پارس JSON با تصحیح خودکار"""
            # تلاش اول - مستقیم
            try:
                return json.loads(json_str), None
            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse JSON from {source_name}: {e}")

                # تلاش دوم - با تصحیح
                try:
                    fixed = try_fix_json(json_str)
                    result = json.loads(fixed)
                    logger.info(f"Successfully parsed JSON from {source_name} after auto-fix")
                    return result, None
                except json.JSONDecodeError as e2:
                    return None, str(e2)

        # 1. ابتدا تلاش برای استخراج از داخل ```json ... ``` یا ``` ... ```
        code_block_match = re.search(r'```(?:json)?\s*([\s\S]*?)```', report_content)
        if code_block_match:
            json_str = code_block_match.group(1).strip()
            report_data, error = attempt_json_parse(json_str, "code block")
            if report_data:
                logger.info(f"Successfully parsed JSON from code block ({len(json_str)} chars)")

        # 2. اگر از code block استخراج نشد، تلاش برای پیدا کردن مستقیم JSON
        if not report_data:
            # پیدا کردن اولین { و آخرین }
            first_brace = report_content.find('{')
            last_brace = report_content.rfind('}')
            if first_brace != -1 and last_brace != -1 and last_brace > first_brace:
                json_str = report_content[first_brace:last_brace+1]
                report_data, error = attempt_json_parse(json_str, "direct extraction")
                if report_data:
                    logger.info(f"Successfully parsed JSON directly ({len(json_str)} chars)")

        # 3. اگر هنوز پارس نشده، محتوای خام را ذخیره کن
        if not report_data:
            logger.error(f"Could not parse JSON from AI response after all attempts. Content preview: {report_content[:500]}")
            report_data = {"raw_content": report_content, "parse_error": True}
        else:
            logger.info(f"Report data parsed successfully. Keys: {list(report_data.keys())}")

        # ====================================
        # 🆕 پردازش نتایج اعتبارسنجی health analysis
        # ====================================
        validated_issues_count = 0
        rejected_issues_count = 0
        new_rejected_archive = []

        # لاگ برای debug
        logger.info(f"🔍 Checking for health_analysis_validation in report_data...")
        logger.info(f"   - validate_health_issues param: {validate_health_issues}")
        logger.info(f"   - 'health_analysis_validation' in report_data: {'health_analysis_validation' in report_data}")
        logger.info(f"   - Health issues sent to AI: {len(health_analysis_issues)}")
        if report_data and not report_data.get("raw_content"):
            logger.info(f"   - Report data keys: {list(report_data.keys())}")

        if validate_health_issues and "health_analysis_validation" in report_data:
            validation_data = report_data["health_analysis_validation"]

            # ذخیره نتایج اعتبارسنجی
            validation_results = {
                "validated_at": datetime.utcnow().isoformat(),
                "validator_model": model_id,
                "total_issues_reviewed": validation_data.get("total_reviewed", 0),
                "validated_count": len(validation_data.get("validated_issues", [])),
                "rejected_count": len(validation_data.get("rejected_issues", [])),
                "validation_summary": validation_data.get("validation_summary", ""),
                "validated_issues": validation_data.get("validated_issues", []),
            }
            project.last_validation_results = json.dumps(validation_results, ensure_ascii=False)
            validated_issues_count = validation_results["validated_count"]
            rejected_issues_count = validation_results["rejected_count"]

            # پردازش ایرادات رد شده و اضافه به آرشیو
            existing_archive = []
            if project.rejected_issues_archive:
                try:
                    existing_archive = json.loads(project.rejected_issues_archive)
                except:
                    pass

            for rejected in validation_data.get("rejected_issues", []):
                archive_entry = {
                    "id": f"rej_{uuid.uuid4().hex[:8]}",
                    "original_issue": rejected.get("original_issue", {}),
                    "source_model": rejected.get("original_issue", {}).get("source_models", ["unknown"])[0] if isinstance(rejected.get("original_issue", {}).get("source_models"), list) else "unknown",
                    "validator_model": model_id,
                    "rejection_reason": rejected.get("rejection_reason", ""),
                    "rejected_at": datetime.utcnow().isoformat(),
                    "validation_score": rejected.get("validation_score", 0),
                }
                new_rejected_archive.append(archive_entry)

            # 🔴 رفع محدودیت - ادغام با آرشیو موجود بدون محدودیت عددی
            combined_archive = new_rejected_archive + existing_archive
            project.rejected_issues_archive = json.dumps(combined_archive, ensure_ascii=False)

            # 🔴 CRITICAL: Commit validation results immediately
            db.commit()
            logger.info(f"✅ Health validation: {validated_issues_count} validated, {rejected_issues_count} rejected - COMMITTED to DB")

            # ====================================
            # 🆕🔴 امتیازدهی به مدل‌های تحلیل سلامت
            # ====================================
            try:
                from ...services.model_profiler import ModelProfiler
                profiler = ModelProfiler()

                # دریافت لیست مدل‌هایی که در تحلیل سلامت شرکت کردند
                source_models_in_validation = set()
                for validated in validation_data.get("validated_issues", []):
                    orig = validated.get("original_issue", {})
                    if isinstance(orig.get("source_models"), list):
                        source_models_in_validation.update(orig["source_models"])
                    elif orig.get("source_model"):
                        source_models_in_validation.add(orig["source_model"])

                for rejected in validation_data.get("rejected_issues", []):
                    orig = rejected.get("original_issue", {})
                    if isinstance(orig.get("source_models"), list):
                        source_models_in_validation.update(orig["source_models"])
                    elif orig.get("source_model"):
                        source_models_in_validation.add(orig["source_model"])

                logger.info(f"🔴 Found {len(source_models_in_validation)} health analysis models to score: {source_models_in_validation}")

                # محاسبه امتیاز برای هر مدل
                for source_model in source_models_in_validation:
                    if not source_model or source_model == "unknown":
                        continue

                    # شمارش ایرادات تایید/رد شده از این مدل
                    correct_from_model = 0
                    false_positives_from_model = 0

                    for validated in validation_data.get("validated_issues", []):
                        orig = validated.get("original_issue", {})
                        models = orig.get("source_models", [orig.get("source_model")])
                        if source_model in models:
                            correct_from_model += 1

                    for rejected in validation_data.get("rejected_issues", []):
                        orig = rejected.get("original_issue", {})
                        models = orig.get("source_models", [orig.get("source_model")])
                        if source_model in models:
                            false_positives_from_model += 1

                    total_from_model = correct_from_model + false_positives_from_model

                    if total_from_model > 0:
                        logger.info(f"📊 Model {source_model}: correct={correct_from_model}, false_positives={false_positives_from_model}")

                        # به‌روزرسانی پروفایل مدل
                        await profiler.update_profile(
                            model_id=source_model,
                            task_type="health_analysis",
                            correct_findings=correct_from_model,
                            total_expected=total_from_model,  # تعداد کل ایراداتی که گزارش کرده
                            false_positives=false_positives_from_model,
                            response_time=0,  # اطلاعات زمان در دسترس نیست
                            tokens_used=0,
                            details={
                                "validated_by": model_id,
                                "project_id": project_id,
                                "validation_date": datetime.utcnow().isoformat(),
                            }
                        )
                        logger.info(f"✅ Updated profile for health analysis model: {source_model}")

            except Exception as prof_error:
                logger.warning(f"⚠️ Could not update model profiles: {prof_error}")

        # 🆕 Fallback: اگر AI بخش health_analysis_validation را برنگرداند (یا JSON پارس نشد)، فیلدها را مستقیماً از health issues بساز
        elif validate_health_issues and health_analysis_issues and (report_data.get("parse_error") or "health_analysis_validation" not in report_data):
            parse_error = report_data.get("parse_error", False)
            has_validation = "health_analysis_validation" in report_data
            logger.warning(f"⚠️ FALLBACK TRIGGERED: parse_error={parse_error}, has_validation={has_validation}")
            logger.warning(f"⚠️ Processing ALL {len(health_analysis_issues)} health issues directly.")

            # 🆕 گروه‌بندی ایرادات براساس severity
            critical_issues = [i for i in health_analysis_issues if i.get("severity") == "critical"]
            high_issues = [i for i in health_analysis_issues if i.get("severity") == "high"]
            medium_issues = [i for i in health_analysis_issues if i.get("severity") == "medium"]
            low_issues = [i for i in health_analysis_issues if i.get("severity") == "low"]

            logger.info(f"📊 Issues breakdown: critical={len(critical_issues)}, high={len(high_issues)}, medium={len(medium_issues)}, low={len(low_issues)}")

            # ذخیره این به عنوان validation results
            validation_results = {
                "validated_at": datetime.utcnow().isoformat(),
                "validator_model": model_id,
                "total_issues_reviewed": len(health_analysis_issues),
                "validated_count": 0,
                "rejected_count": 0,
                "pending_count": 0,  # 🆕 ایرادات در انتظار بررسی
                "validation_summary": "",
                "validated_issues": [],
                "pending_issues": [],  # 🆕 ایرادات medium/low که فیلد ایجاد نشد
                "fallback_mode": True,
            }

            # 🆕 ایجاد فیلد برای تمام critical و high issues (بدون محدودیت 10)
            issues_for_fields = critical_issues + high_issues
            logger.info(f"📝 Creating fields for ALL {len(issues_for_fields)} critical+high issues")

            # اگر critical/high نبود، medium با اولویت بالا رو هم اضافه کن (حداکثر 15)
            if len(issues_for_fields) < 5 and medium_issues:
                additional_medium = medium_issues[:15 - len(issues_for_fields)]
                issues_for_fields.extend(additional_medium)
                logger.info(f"📝 Added {len(additional_medium)} medium issues (total: {len(issues_for_fields)})")

            # ایجاد validated_issues برای فیلدسازی
            for issue in issues_for_fields:
                validation_results["validated_issues"].append({
                    "original_issue": issue,
                    "validation_score": 80 if issue.get("severity") in ["critical", "high"] else 70,
                    "validation_note": "تایید خودکار (Fallback mode) - اولویت بالا",
                    "priority": issue.get("severity"),
                    "create_field": True
                })
                validated_issues_count += 1

            # 🆕 بقیه issues رو به عنوان pending ثبت کن (medium/low که فیلد ایجاد نشد)
            remaining_issues = [i for i in health_analysis_issues if i not in issues_for_fields]
            for issue in remaining_issues:
                validation_results["pending_issues"].append({
                    "original_issue": issue,
                    "validation_score": 50,
                    "validation_note": "در انتظار بررسی - فیلد ایجاد نشد (severity پایین)",
                    "priority": issue.get("severity"),
                    "create_field": False
                })

            validation_results["validated_count"] = validated_issues_count
            validation_results["pending_count"] = len(remaining_issues)
            validation_results["validation_summary"] = f"AI بخش اعتبارسنجی را برنگرداند. از {len(health_analysis_issues)} ایراد: {validated_issues_count} تایید شد (فیلد ایجاد شد)، {len(remaining_issues)} در انتظار بررسی (severity پایین)"
            project.last_validation_results = json.dumps(validation_results, ensure_ascii=False)

            # اضافه کردن به report_data برای پردازش بعدی
            report_data["health_analysis_validation"] = validation_results

            # 🔴 CRITICAL: Commit validation results immediately
            db.commit()
            logger.info(f"✅ Fallback: {validated_issues_count} validated (fields), {len(remaining_issues)} pending - COMMITTED to DB")

        # ====================================
        # 🆕 به‌روزرسانی حالت ایده‌آل جامع
        # ====================================
        if "comprehensive_ideal_state" in report_data:
            ideal_state_data = report_data["comprehensive_ideal_state"]
            comprehensive_ideal = f"""## حالت ایده‌آل پروژه

{ideal_state_data.get('description', '')}

### کمبودهای فعلی:
{chr(10).join(['- ' + d for d in ideal_state_data.get('current_deficiencies', [])])}

### تسک‌های اجرا نشده:
{chr(10).join(['- ' + t for t in ideal_state_data.get('unexecuted_tasks', [])])}

### ساختار سیستم:
{ideal_state_data.get('system_structure', {}).get('overview', '')}

**کامپوننت‌ها:** {', '.join(ideal_state_data.get('system_structure', {}).get('components', []))}

**سیم‌کشی:** {ideal_state_data.get('system_structure', {}).get('wiring', '')}

### نقشه راه یکپارچه:
{ideal_state_data.get('roadmap_integration', '')}
"""
            project.ideal_state = comprehensive_ideal
            logger.info(f"Updated comprehensive ideal state for project {project_id}")

        # مدیریت هوشمند فیلدها
        created_fields = []
        archived_count = 0
        merged_count = 0
        updated_count = 0

        if auto_create_fields:
            # 1. پردازش field_management - بایگانی، ادغام، به‌روزرسانی
            field_mgmt = report_data.get("field_management", {})

            # بایگانی فیلدها
            for field_id in field_mgmt.get("fields_to_archive", []):
                for field in existing_fields:
                    if field.get("id") == field_id and not field.get("archived"):
                        field["archived"] = True
                        field["archived_at"] = datetime.utcnow().isoformat()
                        field["archived_reason"] = "report review"
                        archived_count += 1

            # ادغام فیلدها
            for merge_info in field_mgmt.get("fields_to_merge", []):
                source_ids = merge_info.get("source_ids", [])
                if len(source_ids) >= 2:
                    merged_attachments = []
                    for sid in source_ids:
                        for field in existing_fields:
                            if field.get("id") == sid:
                                merged_attachments.extend(field.get("attachments", []))
                                field["archived"] = True
                                field["archived_at"] = datetime.utcnow().isoformat()
                                field["archived_reason"] = "merged"

                    merged_field = {
                        "id": str(uuid.uuid4()),
                        "name": merge_info.get("merged_name", "فیلد ادغام‌شده"),
                        "value": merge_info.get("merged_value", ""),
                        "target_models": ["claude"],
                        "action_type": "display",
                        "field_type": "permanent",
                        "priority": 5,
                        "attachments": merged_attachments,
                        "trigger": {"enabled": False, "interval_minutes": 60, "interval_type": "minutes"},
                        "created_from_report": True,
                        "merged_from": source_ids,
                    }
                    existing_fields.append(merged_field)
                    merged_count += 1

            # به‌روزرسانی فیلدها
            for update_info in field_mgmt.get("fields_to_update", []):
                field_id = update_info.get("id")
                for field in existing_fields:
                    if field.get("id") == field_id and not field.get("archived"):
                        if update_info.get("new_value"):
                            field["value"] = update_info["new_value"]
                        if update_info.get("new_priority"):
                            field["priority"] = update_info["new_priority"]
                        field["updated_at"] = datetime.utcnow().isoformat()
                        updated_count += 1

            # 🔴 تایید فیلدهای pending (ایجاد شده توسط AI query یا بخش‌های دیگر)
            approved_count = 0
            for field_id in field_mgmt.get("fields_to_approve", []):
                for field in existing_fields:
                    if field.get("id") == field_id and not field.get("archived"):
                        field["engineering_approval"] = {
                            "approved": True,
                            "approved_at": datetime.utcnow().isoformat(),
                            "approved_by": model_id,
                            "approval_type": "engineering_review"
                        }
                        field["validation_marker"] = "validated"
                        field["needs_approval"] = False
                        approved_count += 1
                        logger.info(f"Approved pending field: {field.get('name')}")

            # 🔴 رد و بایگانی فیلدهای pending که مورد نیاز نیستند
            rejected_field_count = 0
            for field_id in field_mgmt.get("fields_to_reject", []):
                for field in existing_fields:
                    if field.get("id") == field_id and not field.get("archived"):
                        field["archived"] = True
                        field["archived_at"] = datetime.utcnow().isoformat()
                        field["archived_reason"] = "rejected_by_engineering_report"
                        rejected_field_count += 1
                        logger.info(f"Rejected pending field: {field.get('name')}")

            # 2. 🆕 تولید فیلدهای جدید از validated issues (با مارکر اعتبارسنجی)
            if validate_health_issues and "health_analysis_validation" in report_data:
                validated_issues = report_data["health_analysis_validation"].get("validated_issues", [])
                logger.info(f"Creating fields from {len(validated_issues)} validated health issues")

                # 🔴 دریافت سوابق اجرا برای جلوگیری از ایجاد مجدد
                executed_issues = set()
                for field in existing_fields:
                    if field.get("executed") or field.get("archived"):
                        # استخراج اطلاعات اصلی از فیلد برای تشخیص تکراری
                        orig = field.get("original_issue", {})
                        if orig:
                            issue_key = f"{orig.get('file', '')}:{orig.get('type', '')}:{orig.get('line', '')}"
                            executed_issues.add(issue_key.lower())
                        # همچنین نام فیلد را به عنوان کلید ذخیره کن
                        executed_issues.add(field.get("name", "").lower())

                logger.info(f"Found {len(executed_issues)} previously executed/archived issues")

                for issue in validated_issues:
                    if not issue.get("create_field", True):
                        continue

                    original = issue.get("original_issue", {})
                    field_name = f"✅ [تایید شده] {original.get('type', 'issue')}: {original.get('file', 'unknown')}"

                    # 🔴 بررسی وجود فیلد مشابه - شامل بایگانی شده‌ها
                    # ۱. بررسی با نام
                    existing_by_name = any(
                        f.get("name", "").lower() == field_name.lower()
                        for f in existing_fields  # همه فیلدها، نه فقط غیر بایگانی
                    )

                    # ۲. بررسی با کلید یکتا (فایل + نوع + خط)
                    issue_key = f"{original.get('file', '')}:{original.get('type', '')}:{original.get('line', '')}"
                    existing_by_key = issue_key.lower() in executed_issues

                    # ۳. بررسی در سوابق اجرا
                    already_executed = field_name.lower() in executed_issues

                    if existing_by_name or existing_by_key or already_executed:
                        logger.info(f"Skipping already processed issue: {field_name}")
                        continue

                    # 🔴 این کد باید بعد از continue باشد، نه داخل if
                    priority_map = {"critical": 1, "high": 2, "medium": 5, "low": 7}
                    priority = priority_map.get(issue.get("priority", original.get("severity", "medium")), 5)

                    # تعیین هوشمند action_type براساس نوع مشکل و محتوای پیام
                    target_file = original.get("file")
                    issue_type = original.get("type", "").lower()
                    issue_message = original.get("message", "").lower()

                    # تعیین نوع اقدام براساس کلیدواژه‌ها
                    field_action_type = "display"  # پیش‌فرض
                    trigger_enabled = False
                    trigger_interval = 60
                    trigger_type = "minutes"

                    # انواع مشکلاتی که نیاز به تغییر کد دارند
                    code_change_types = ["security", "bug", "quality", "performance", "error", "warning", "vulnerability", "fix", "اصلاح"]
                    needs_code_change = any(t in issue_type for t in code_change_types)

                    # کلیدواژه‌های حذف فایل
                    delete_keywords = ["delete", "remove", "حذف", "unused", "deprecated", "غیرضروری"]
                    needs_delete = any(kw in issue_message for kw in delete_keywords)

                    # کلیدواژه‌های ایجاد فایل جدید
                    create_keywords = ["create", "add file", "new file", "missing", "ایجاد", "اضافه کن", "فایل جدید"]
                    needs_create = any(kw in issue_message for kw in create_keywords)

                    # کلیدواژه‌های جابجایی/تغییر نام
                    move_keywords = ["move", "rename", "relocate", "جابجا", "تغییر نام", "انتقال"]
                    needs_move = any(kw in issue_message for kw in move_keywords)

                    # کلیدواژه‌های نیازمند اجرای دوره‌ای
                    periodic_keywords = ["monitor", "check regularly", "periodic", "backup", "sync", "مانیتور", "بررسی دوره‌ای", "پشتیبان"]
                    needs_periodic = any(kw in issue_message for kw in periodic_keywords)

                    # تعیین action_type
                    if target_file:
                        if needs_delete:
                            field_action_type = "github_delete"
                        elif needs_move:
                            field_action_type = "github_move"
                        elif needs_create:
                            field_action_type = "github_create"
                        elif needs_code_change:
                            field_action_type = "github_commit"

                    # تنظیم تریگر برای موارد دوره‌ای
                    if needs_periodic:
                        trigger_enabled = True
                        trigger_interval = 24 * 60  # روزانه
                        trigger_type = "minutes"

                    new_field = {
                        "id": str(uuid.uuid4()),
                        "name": field_name,
                        "value": f"""## ایراد تایید شده توسط {model_id}

**فایل:** {original.get('file', 'نامشخص')}
**نوع:** {original.get('type', 'نامشخص')}
**شدت:** {original.get('severity', 'نامشخص')}
**خط:** {original.get('line', 'نامشخص')}
**نوع اقدام:** {field_action_type}

### پیام:
{original.get('message', '')}

### یادداشت اعتبارسنجی:
{issue.get('validation_note', '')}

### امتیاز اعتبارسنجی: {issue.get('validation_score', 0)}/100

---
لطفاً این مشکل را بررسی و رفع کنید. کد اصلاح شده را تولید کن.
""",
                        "target_models": ["claude"],
                        "action_type": field_action_type,
                        "target_path": target_file,
                        "archive_after_run": not trigger_enabled,  # اگر تریگر فعال است، بایگانی نشود
                        "deploy_after_commit": field_action_type.startswith("github_"),
                        "field_type": "permanent" if trigger_enabled else "temporary",
                        "priority": priority,
                        "attachments": [],
                        "trigger": {
                            "enabled": trigger_enabled,
                            "interval_minutes": trigger_interval,
                            "interval_type": trigger_type
                        },
                        "created_from_report": True,
                        "validation_marker": "validated",
                        "validation_score": issue.get("validation_score", 0),
                        "validator_model": model_id,
                        "original_issue": original,
                        # 🔴 تاییدیه گزارش مهندسی - فقط فیلدهای دارای این تاییدیه قابل اجرا هستند
                        "engineering_approval": {
                            "approved": True,
                            "approved_at": datetime.utcnow().isoformat(),
                            "approved_by": model_id,
                            "approval_type": "health_validation"
                        }
                    }
                    existing_fields.append(new_field)
                    created_fields.append(new_field["name"])

                    # 🔴 بایگانی ایراد اصلی در لیست issues_found - 🆕 منطق بهبود یافته
                    try:
                        issues_found = []
                        if project.issues_found:
                            issues_found = json.loads(project.issues_found)

                        # استخراج اطلاعات برای تطبیق
                        orig_file = (original.get("file", original.get("file_path", "")) or "").lower()
                        orig_type = (original.get("type", "") or "").lower()
                        orig_message = (original.get("message", original.get("description", "")) or "").lower().strip()
                        orig_line = original.get("line") or original.get("start_line")
                        orig_stable_id = original.get("stable_id")
                        orig_id = original.get("id")

                        # استخراج کلمات کلیدی از پیام برای تطبیق fuzzy
                        orig_words = set(w for w in orig_message.split() if len(w) > 3)

                        # پیدا کردن و بایگانی ایراد با روش‌های مختلف
                        archived_count = 0
                        for stored_issue in issues_found:
                            if stored_issue.get("archived"):
                                continue

                            stored_file = (stored_issue.get("file", stored_issue.get("file_path", "")) or "").lower()
                            stored_type = (stored_issue.get("type", "") or "").lower()
                            stored_message = (stored_issue.get("message", stored_issue.get("description", "")) or "").lower().strip()
                            stored_line = stored_issue.get("line") or stored_issue.get("start_line")
                            stored_words = set(w for w in stored_message.split() if len(w) > 3)

                            match_found = False

                            # روش 1: تطبیق با stable_id
                            if orig_stable_id and stored_issue.get("stable_id") == orig_stable_id:
                                match_found = True

                            # روش 2: تطبیق با id
                            elif orig_id and stored_issue.get("id") == orig_id:
                                match_found = True

                            # روش 3: تطبیق دقیق فایل + نوع + خط
                            elif orig_file and stored_file and orig_file == stored_file:
                                if orig_type == stored_type:
                                    match_found = True
                                elif orig_line and stored_line and orig_line == stored_line:
                                    match_found = True

                            # روش 4: تطبیق فایل (substring) + تطبیق fuzzy پیام
                            elif orig_file and stored_file:
                                file_match = (orig_file in stored_file or stored_file in orig_file or
                                             orig_file.split('/')[-1] == stored_file.split('/')[-1])
                                if file_match and orig_words and stored_words:
                                    common_words = orig_words.intersection(stored_words)
                                    # اگر حداقل 30% کلمات مشترک باشند
                                    if len(common_words) >= max(1, min(len(orig_words), len(stored_words)) * 0.3):
                                        match_found = True

                            # روش 5: تطبیق پیام (برای ایراداتی که فایل ندارند)
                            elif not orig_file and not stored_file and orig_message and stored_message:
                                if orig_message in stored_message or stored_message in orig_message:
                                    match_found = True

                            if match_found:
                                stored_issue["archived"] = True
                                stored_issue["archived_at"] = datetime.utcnow().isoformat()
                                stored_issue["archived_reason"] = "converted_to_field"
                                stored_issue["field_id"] = new_field["id"]
                                archived_count += 1
                                logger.info(f"Archived issue: {stored_file}:{stored_type} -> {new_field['name']}")

                        project.issues_found = json.dumps(issues_found, ensure_ascii=False)
                        if archived_count > 0:
                            logger.info(f"✅ Archived {archived_count} matching issues for field: {new_field['name']}")
                        else:
                            logger.warning(f"⚠️ No matching issues found for: {orig_file}:{orig_type}")
                    except Exception as e:
                        logger.warning(f"Could not archive issue: {e}")

            # 3. تولید فیلدهای جدید از roadmap
            roadmap = report_data.get("roadmap", {})

            for phase_name, phase_tasks in [("immediate", roadmap.get("immediate", [])), ("short_term", roadmap.get("short_term", []))]:
                for task in phase_tasks[:5]:
                    field_name = task.get("task", "تسک جدید")
                    field_value = task.get("description", "")
                    action_type = task.get("action_type", "display")
                    target_path = task.get("target_path")
                    field_type = task.get("field_type", "temporary")
                    priority = task.get("priority", 5)

                    # بررسی وجود فیلد مشابه
                    existing = any(
                        f.get("name", "").lower() == field_name.lower()
                        for f in existing_fields if not f.get("archived")
                    )

                    if not existing and field_value:
                        new_field = {
                            "id": str(uuid.uuid4()),
                            "name": f"[{phase_name}] {field_name}",
                            "value": field_value,
                            "target_models": ["claude"],
                            "action_type": action_type,
                            "target_path": target_path,
                            "archive_after_run": field_type == "temporary",
                            "deploy_after_commit": action_type == "github_commit",
                            "field_type": field_type,
                            "priority": priority,
                            "attachments": [],
                            "trigger": {"enabled": False, "interval_minutes": 60, "interval_type": "minutes"},
                            "created_from_report": True,
                            # 🔴 تاییدیه گزارش مهندسی
                            "engineering_approval": {
                                "approved": True,
                                "approved_at": datetime.utcnow().isoformat(),
                                "approved_by": model_id,
                                "approval_type": "roadmap_task"
                            }
                        }
                        existing_fields.append(new_field)
                        created_fields.append(new_field["name"])

            # 4. 🆕 تولید فیلدهای عملیاتی از journal_analysis
            if "journal_analysis" in report_data:
                journal_data = report_data["journal_analysis"]
                findings = journal_data.get("findings", [])
                logger.info(f"Processing {len(findings)} journal findings for actionable fields")

                for finding in findings:
                    if not finding.get("create_actionable_field", False):
                        continue

                    suggested = finding.get("suggested_field", {})
                    if not suggested:
                        continue

                    field_name = suggested.get("name", f"[ژورنال] {finding.get('summary', 'تسک')}")
                    field_value = suggested.get("value", finding.get("summary", ""))
                    action_type = suggested.get("action_type", "github_commit")  # 🔴 پیش‌فرض عملیاتی
                    target_path = suggested.get("target_path")
                    priority = suggested.get("priority", 3)

                    # اگر action_type نمایشی بود، به github_commit تغییر بده
                    if action_type == "display" and target_path:
                        action_type = "github_commit"
                        logger.info(f"Changed action_type from display to github_commit for: {field_name}")

                    # بررسی تکراری
                    existing = any(
                        f.get("name", "").lower() == field_name.lower()
                        for f in existing_fields if not f.get("archived")
                    )

                    if not existing and field_value:
                        new_field = {
                            "id": str(uuid.uuid4()),
                            "name": f"📋 {field_name}",
                            "value": f"""## یافته از ژورنال

**نوع یافته:** {finding.get('finding_type', 'unknown')}
**شدت:** {finding.get('severity', 'medium')}
**منبع:** ردیف ژورنال {finding.get('journal_entry_id', 'نامشخص')} - فیلد "{finding.get('field_name', 'نامشخص')}"

### خلاصه یافته:
{finding.get('summary', '')}

### دستور اجرا:
{field_value}

---
این فیلد عملیاتی است و باید کد تولید/اصلاح شود.
""",
                            "target_models": ["claude"],
                            "action_type": action_type,
                            "target_path": target_path,
                            "archive_after_run": True,
                            "deploy_after_commit": action_type.startswith("github_"),
                            "field_type": "temporary",
                            "priority": priority,
                            "attachments": [],
                            "trigger": {"enabled": False, "interval_minutes": 60, "interval_type": "minutes"},
                            "created_from_report": True,
                            "source": "journal_analysis",
                            "journal_entry_id": finding.get("journal_entry_id"),
                            # 🔴 تاییدیه گزارش مهندسی
                            "engineering_approval": {
                                "approved": True,
                                "approved_at": datetime.utcnow().isoformat(),
                                "approved_by": model_id,
                                "approval_type": "journal_finding"
                            }
                        }
                        existing_fields.append(new_field)
                        created_fields.append(new_field["name"])
                        logger.info(f"Created actionable field from journal: {new_field['name']} (action: {action_type})")

            # 5. 🔴 به‌روزرسانی نقشه راه با چک‌باکس‌ها
            roadmap_updated = False
            if "roadmap_status_updates" in report_data:
                roadmap_updates = report_data["roadmap_status_updates"]
                current_roadmap = project.roadmap_content or ""

                for update in roadmap_updates:
                    item_name = update.get("item", "")
                    completed = update.get("completed", False)

                    if completed:
                        # ✅ تیک سبز برای موارد انجام شده
                        # تبدیل [ ] به [x] یا اضافه کردن ✅
                        if f"[ ] {item_name}" in current_roadmap:
                            current_roadmap = current_roadmap.replace(f"[ ] {item_name}", f"[x] {item_name} ✅")
                            roadmap_updated = True
                        elif item_name in current_roadmap and "[x]" not in current_roadmap.split(item_name)[0][-10:]:
                            # اگر چک‌باکس ندارد، ✅ اضافه کن
                            current_roadmap = current_roadmap.replace(item_name, f"{item_name} ✅")
                            roadmap_updated = True

                    elif update.get("create_field"):
                        # ایجاد فیلد برای آیتم‌های انجام نشده
                        field_details = update.get("field_details", {})
                        new_field = {
                            "id": str(uuid.uuid4()),
                            "name": f"[نقشه راه] {field_details.get('name', item_name)}",
                            "value": field_details.get("value", f"تسک نقشه راه: {item_name}"),
                            "target_models": ["claude"],
                            "action_type": field_details.get("action_type", "github_commit"),
                            "target_path": field_details.get("target_path"),
                            "archive_after_run": True,
                            "field_type": "temporary",
                            "priority": field_details.get("priority", 3),
                            "attachments": [],
                            "trigger": {"enabled": False, "interval_minutes": 60, "interval_type": "minutes"},
                            "created_from_report": True,
                            "source": "roadmap",
                            "roadmap_item": item_name,
                            "engineering_approval": {
                                "approved": True,
                                "approved_at": datetime.utcnow().isoformat(),
                                "approved_by": model_id,
                                "approval_type": "roadmap_item"
                            }
                        }
                        existing_fields.append(new_field)
                        created_fields.append(new_field["name"])
                        logger.info(f"Created field from roadmap: {new_field['name']}")

                if roadmap_updated:
                    project.roadmap_content = current_roadmap
                    logger.info(f"Roadmap updated with {len(roadmap_updates)} status changes")

            # 6. مرتب‌سازی براساس اولویت و ذخیره
            active = [f for f in existing_fields if not f.get("archived")]
            archived = [f for f in existing_fields if f.get("archived")]
            active.sort(key=lambda x: int(x.get("priority", 5)) if str(x.get("priority", 5)).isdigit() else 5)
            existing_fields = active + archived

            if created_fields or archived_count or merged_count or updated_count or roadmap_updated:
                project.dynamic_fields = json.dumps(existing_fields, ensure_ascii=False)
                db.commit()
                logger.info(f"Field management: created={len(created_fields)}, archived={archived_count}, merged={merged_count}, updated={updated_count}")

        # ایجاد گزارش
        total_tokens = sum(log.tokens_used for log in logs) + (response.tokens_used or 0)
        models_used = list(set(log.model_id for log in logs))
        models_used.append(model_id)

        report = Report(
            id=f"eng_report_{uuid.uuid4().hex[:12]}",
            project_id=project_id,
            report_type="engineering",
            title=f"گزارش مهندسی - {project.name}",
            content=json.dumps(report_data, ensure_ascii=False, indent=2),
            summary=report_data.get("executive_summary", f"گزارش مهندسی با تحلیل {len(files)} فایل"),
            total_activities=len(logs),
            total_tokens=total_tokens,
            models_used=json.dumps(list(set(models_used))),
            period_start=since,
            period_end=datetime.utcnow(),
            created_at=datetime.utcnow(),
            generated_by=model_id,
        )

        db.add(report)
        db.commit()

        # 🆕 ثبت فعالیت در ژورنال
        activity_log = ActivityLog(
            id=f"log_{uuid.uuid4().hex[:12]}",
            project_id=project_id,
            model_id=model_id,
            model_provider="anthropic" if "claude" in model_id.lower() else "openai",
            activity_type="engineering_report",
            prompt=f"تولید گزارش مهندسی برای {days} روز اخیر",
            response=report_data.get("executive_summary", "")[:500] if isinstance(report_data, dict) else None,
            tokens_used=response.tokens_used or 0,
            latency_ms=int((datetime.utcnow() - since).total_seconds() * 1000) if since else 0,
            success=True,
            field_id=None,
            field_name=f"گزارش مهندسی - {report.id}",
            extra_data=json.dumps({
                "report_id": report.id,
                "files_analyzed": len(files),
                "health_issues_reviewed": len(health_analysis_issues),
                "validated_count": validated_issues_count,
                "rejected_count": rejected_issues_count,
                "fields_created": len(created_fields),
                "project_health_score": report_data.get("project_health", {}).get("score") if isinstance(report_data, dict) else None,
            }, ensure_ascii=False),
            created_at=datetime.utcnow(),
        )
        db.add(activity_log)
        db.commit()

        return {
            "success": True,
            "report_id": report.id,
            "report_type": "engineering",
            "message": f"گزارش مهندسی جامع تولید شد",
            # 🔴 اطلاعات مدل استفاده شده (fallback handling)
            "model_used": model_id,
            "original_model_requested": original_model_id,
            "used_fallback": used_fallback,
            "fields_created": created_fields,
            "fields_count": len(created_fields),
            "fields_archived": archived_count,
            "fields_merged": merged_count,
            "fields_updated": updated_count,
            "project_health_score": report_data.get("project_health", {}).get("score"),
            "bugs_found": len(report_data.get("bugs_and_issues", [])),
            "recommendations_count": len(report_data.get("recommendations", [])),
            # 🆕 نتایج اعتبارسنجی
            "validation_results": {
                "issues_reviewed": validated_issues_count + rejected_issues_count,
                "validated_count": validated_issues_count,
                "rejected_count": rejected_issues_count,
                "fields_created_from_validation": len([f for f in created_fields if "✅" in f]),
            } if validate_health_issues else None,
            "ideal_state_updated": "comprehensive_ideal_state" in report_data,
            # 🆕 نتایج بررسی ژورنال
            "journal_analysis": {
                "entries_reviewed": len(journal_entries_for_review),
                "entries_with_findings": len(journal_reviews),
                "fields_created_from_journal": len([f for f in created_fields if "📋" in f]),
            } if journal_entries_for_review else None,
            # 🔴 نتایج بررسی عمیق GitHub
            "github_deep_inspection": github_inspection_result,
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"خطا در تولید گزارش: {str(e)}",
            # 🔴 اطلاعات مدل برای debug
            "model_attempted": model_id,
            "original_model_requested": original_model_id,
            "used_fallback": used_fallback,
        }


# ===================== گزارش مهندسی با نوار پیشرفت =====================

@router.post("/{project_id}/reports/generate-engineering-stream")
async def generate_engineering_report_stream(
    project_id: str,
    days: int = Query(7, ge=1, le=30),
    model_id: str = Query(None),  # برای backward compatibility
    model_ids: str = Query(None),  # 🆕 چند مدل با کاما جدا شده
    depth: str = Query("standard"),  # 🆕 quick, standard, deep
    auto_create_fields: bool = Query(True),
    validate_health_issues: bool = Query(True),
    db: Session = Depends(get_db)
):
    """
    🔴 گزارش مهندسی با نوار پیشرفت (Streaming)

    🆕 قابلیت‌های جدید:
    - model_ids: چند مدل با کاما جدا شده (مثلاً "claude,gpt-4o")
    - depth: عمق تحلیل
        - quick: بررسی سریع (1-2 دقیقه)
        - standard: تحلیل متوسط (3-5 دقیقه)
        - deep: تحلیل عمیق فایل به فایل (10-20 دقیقه)
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.ai_base import Message
    from ...models.project import ProjectFile
    import logging
    logger = logging.getLogger(__name__)

    # پردازش مدل‌ها
    selected_models = []
    if model_ids:
        selected_models = [m.strip() for m in model_ids.split(",") if m.strip()]
    elif model_id:
        selected_models = [model_id]
    if not selected_models:
        selected_models = ["claude"]  # پیش‌فرض

    # 🔴 جلوگیری از دور باطل - بررسی گزارش تکراری
    if check_cycle_prevention(
        db=db,
        project_id=project_id,
        activity_type="engineering_report",
        minutes_threshold=5  # حداقل 5 دقیقه بین گزارش‌ها
    ) == False:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=429,
            content={"error": "گزارش مهندسی اخیراً تولید شده است. لطفاً چند دقیقه صبر کنید."}
        )

    # تعداد مراحل بر اساس عمق - 🔴 بدون محدودیت فایل (تنظیمات جداگانه)
    depth_config = {
        "quick": {"total_steps": 4, "ai_calls": 1, "delay_factor": 0.2, "file_delay": 0.02, "use_4step": False},
        "standard": {"total_steps": 8, "ai_calls": 2, "delay_factor": 1.0, "file_delay": 0.1, "use_4step": False},
        "deep": {"total_steps": 20, "ai_calls": 8, "delay_factor": 5.0, "file_delay": 0.5, "use_4step": True}  # 🔴 فرآیند 4 مرحله‌ای واقعی
    }
    config = depth_config.get(depth, depth_config["standard"])
    delay_factor = config.get("delay_factor", 0.5)
    file_delay = config.get("file_delay", 0.05)
    use_4step = config.get("use_4step", False)  # 🆕 در deep از فرآیند 4 مرحله‌ای استفاده شود

    async def progress_generator():
        """Generator برای ارسال پیشرفت"""
        try:
            total_steps = config["total_steps"]

            # مرحله 1: بررسی پروژه
            yield f"data: {json.dumps({'step': 1, 'total': total_steps, 'message': '🔍 بررسی پروژه...', 'progress': 5}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(delay_factor)  # 🔴 استفاده از delay_factor

            project = db.query(Project).filter(Project.id == project_id).first()
            if not project:
                yield f"data: {json.dumps({'error': 'پروژه یافت نشد'}, ensure_ascii=False)}\n\n"
                return

            # مرحله 2: دریافت فایل‌ها
            yield f"data: {json.dumps({'step': 2, 'total': total_steps, 'message': '📂 دریافت فایل‌های پروژه...', 'progress': 10}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(delay_factor)  # 🔴 استفاده از delay_factor

            files = db.query(ProjectFile).filter(ProjectFile.project_id == project_id).all()
            # 🔴 بدون محدودیت - همه فایل‌ها تحلیل می‌شوند
            files_to_analyze = files

            # 🆕 در حالت deep، هر فایل را گزارش کن
            if depth == "deep":
                for i, f in enumerate(files_to_analyze):
                    progress = 10 + int((i / max(len(files_to_analyze), 1)) * 15)
                    yield f"data: {json.dumps({'step': 2, 'message': f'📄 [{i+1}/{len(files_to_analyze)}] {f.file_path}', 'progress': progress}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(file_delay)  # 🔴 0.3 ثانیه برای هر فایل در deep
            else:
                for i, f in enumerate(files_to_analyze[:10]):
                    yield f"data: {json.dumps({'step': 2, 'message': f'📄 بررسی: {f.file_path}', 'progress': 10 + (i * 1)}, ensure_ascii=False)}\n\n"
                    await asyncio.sleep(file_delay)

            # مرحله 3: دریافت فیلدها
            yield f"data: {json.dumps({'step': 3, 'total': total_steps, 'message': '📋 بررسی فیلدهای پویا...', 'progress': 30}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(delay_factor)  # 🔴 استفاده از delay_factor

            existing_fields = []
            try:
                if project.dynamic_fields:
                    existing_fields = json.loads(project.dynamic_fields)
            except:
                pass

            pending_fields = [f for f in existing_fields if not f.get("archived") and not f.get("engineering_approval")]
            yield f"data: {json.dumps({'step': 3, 'message': f'🔴 {len(pending_fields)} فیلد pending برای اعتبارسنجی', 'progress': 35}, ensure_ascii=False)}\n\n"

            # مرحله 4: بررسی health issues
            yield f"data: {json.dumps({'step': 4, 'total': total_steps, 'message': '🔍 بررسی ایرادات سلامت...', 'progress': 40}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(delay_factor)  # 🔴 استفاده از delay_factor

            health_issues = []
            if project.issues_found:
                try:
                    health_issues = json.loads(project.issues_found)
                except:
                    pass

            # 🔴 فیلتر ایرادات غیربایگانی
            active_issues = [i for i in health_issues if not i.get("archived")]
            yield f"data: {json.dumps({'step': 4, 'message': f'⚠️ {len(active_issues)} ایراد فعال برای اعتبارسنجی (کل: {len(health_issues)})', 'progress': 45}, ensure_ascii=False)}\n\n"

            # 🆕 مراحل اضافی برای عمق deep
            if depth == "deep" and len(selected_models) > 1:
                # مرحله 5-8: تحلیل توسط هر مدل
                model_results = []
                step = 5
                for model in selected_models:
                    yield f"data: {json.dumps({'step': step, 'total': total_steps, 'message': f'🧠 تحلیل توسط {model}...', 'progress': 50 + (step - 5) * 10}, ensure_ascii=False)}\n\n"

                    try:
                        # فراخوانی واقعی AI برای هر مدل
                        ai_manager = get_ai_manager()
                        file_summary = "\n".join([f"- {f.file_path}" for f in files_to_analyze[:30]])

                        # تحلیل ساختار با این مدل خاص
                        analysis_prompt = f"""
تحلیل مهندسی پروژه {project.name}:

فایل‌ها ({len(files_to_analyze)} فایل):
{file_summary}

وظیفه: تحلیل کیفیت کد، امنیت، و ساختار. حداکثر 5 ایراد مهم را شناسایی کن.

پاسخ را به فرمت JSON بده:
{{"issues": [{{"type": "", "severity": "high/medium/low", "file": "", "description": ""}}], "score": 0-100}}
"""
                        response = await ai_manager.generate(
                            model_id=model,
                            messages=[Message(role="user", content=analysis_prompt)],
                            max_tokens=2000,
                            temperature=0.3
                        )
                        if response.content:
                            model_results.append({"model": model, "response": response.content[:500]})
                            yield f"data: {json.dumps({'step': step, 'message': f'✅ {model}: تحلیل کامل شد', 'progress': 50 + (step - 4) * 10}, ensure_ascii=False)}\n\n"

                    except Exception as me:
                        yield f"data: {json.dumps({'step': step, 'message': f'⚠️ {model}: خطا - {str(me)[:50]}', 'progress': 50 + (step - 4) * 10}, ensure_ascii=False)}\n\n"

                    step += 1
                    await asyncio.sleep(delay_factor * 2)  # 🔴 6 ثانیه بین هر مدل در deep

            # مرحله قبل از آخر: بررسی نقشه راه
            step = total_steps - 2
            yield f"data: {json.dumps({'step': step, 'total': total_steps, 'message': '🗺️ بررسی نقشه راه و حالت ایده‌آل...', 'progress': 70}, ensure_ascii=False)}\n\n"
            await asyncio.sleep(delay_factor)  # 🔴 استفاده از delay_factor

            # 🔴 در حالت deep از فرآیند 4 مرحله‌ای استفاده شود
            if use_4step:
                # مرحله 1: اعتبارسنجی فیلدها
                step = total_steps - 5
                msg1_start = json.dumps({'step': step, 'total': total_steps, 'message': '🔬 مرحله ۱: اعتبارسنجی فیلدهای موجود...', 'progress': 60}, ensure_ascii=False)
                yield f"data: {msg1_start}\n\n"
                await asyncio.sleep(delay_factor)
                try:
                    step1_result = await engineering_step1_validate_fields(project_id, selected_models[0], depth, db)
                    validated_count = step1_result.get('validated_count', 0)
                    msg1_done = json.dumps({'step': step, 'message': f'✅ مرحله ۱ تکمیل: {validated_count} فیلد بررسی شد', 'progress': 65}, ensure_ascii=False)
                    yield f"data: {msg1_done}\n\n"
                except Exception as e:
                    err_msg = str(e)[:50]
                    msg1_err = json.dumps({'step': step, 'message': f'⚠️ مرحله ۱: {err_msg}', 'progress': 65}, ensure_ascii=False)
                    yield f"data: {msg1_err}\n\n"
                await asyncio.sleep(delay_factor)

                # مرحله 2: تبدیل ایرادات سلامت به فیلد
                step = total_steps - 4
                msg2_start = json.dumps({'step': step, 'total': total_steps, 'message': '🔄 مرحله ۲: تبدیل ایرادات سلامت به فیلد...', 'progress': 70}, ensure_ascii=False)
                yield f"data: {msg2_start}\n\n"
                await asyncio.sleep(delay_factor)
                try:
                    step2_result = await engineering_step2_health_to_fields(project_id, selected_models[0], depth, db)
                    created = step2_result.get('created_count', 0)
                    archived = step2_result.get('archived_count', 0)
                    msg2_done = json.dumps({'step': step, 'message': f'✅ مرحله ۲: {created} فیلد ایجاد، {archived} ایراد بایگانی', 'progress': 75}, ensure_ascii=False)
                    yield f"data: {msg2_done}\n\n"
                except Exception as e:
                    err_msg = str(e)[:50]
                    msg2_err = json.dumps({'step': step, 'message': f'⚠️ مرحله ۲: {err_msg}', 'progress': 75}, ensure_ascii=False)
                    yield f"data: {msg2_err}\n\n"
                await asyncio.sleep(delay_factor)

                # مرحله 3: ارزیابی مدل‌ها
                step = total_steps - 3
                msg3_start = json.dumps({'step': step, 'total': total_steps, 'message': '📊 مرحله ۳: ارزیابی عملکرد مدل‌ها...', 'progress': 80}, ensure_ascii=False)
                yield f"data: {msg3_start}\n\n"
                await asyncio.sleep(delay_factor)
                try:
                    step3_result = await engineering_step3_evaluate_models(project_id, selected_models[0], depth, db)
                    msg3_done = json.dumps({'step': step, 'message': '✅ مرحله ۳: ارزیابی مدل‌ها تکمیل شد', 'progress': 85}, ensure_ascii=False)
                    yield f"data: {msg3_done}\n\n"
                except Exception as e:
                    err_msg = str(e)[:50]
                    msg3_err = json.dumps({'step': step, 'message': f'⚠️ مرحله ۳: {err_msg}', 'progress': 85}, ensure_ascii=False)
                    yield f"data: {msg3_err}\n\n"
                await asyncio.sleep(delay_factor)

                # مرحله 4: به‌روزرسانی نقشه راه
                step = total_steps - 2
                msg4_start = json.dumps({'step': step, 'total': total_steps, 'message': '🗺️ مرحله ۴: به‌روزرسانی نقشه راه...', 'progress': 90}, ensure_ascii=False)
                yield f"data: {msg4_start}\n\n"
                await asyncio.sleep(delay_factor)
                try:
                    step4_result = await engineering_step4_update_roadmap(project_id, selected_models[0], db)
                    msg4_done = json.dumps({'step': step, 'message': '✅ مرحله ۴: نقشه راه به‌روزرسانی شد', 'progress': 95}, ensure_ascii=False)
                    yield f"data: {msg4_done}\n\n"
                except Exception as e:
                    err_msg = str(e)[:50]
                    msg4_err = json.dumps({'step': step, 'message': f'⚠️ مرحله ۴: {err_msg}', 'progress': 95}, ensure_ascii=False)
                    yield f"data: {msg4_err}\n\n"
                await asyncio.sleep(delay_factor)

                # 🟢 اصلاح شده: فقط ایرادات تبدیل‌شده به فیلد بایگانی شوند
                # ایرادات با archived_reason="converted_to_field" قبلاً در generate_engineering_report بایگانی شده‌اند
                # بقیه ایرادات باید فعال بمانند تا کاربر بتواند آنها را دستی بررسی کند
                try:
                    project_fresh = db.query(Project).filter(Project.id == project_id).first()
                    if project_fresh and project_fresh.issues_found:
                        all_issues = json.loads(project_fresh.issues_found)
                        # فقط شمارش ایرادات تبدیل‌شده (که قبلاً بایگانی شده‌اند)
                        converted_count = sum(1 for i in all_issues if i.get("archived_reason") == "converted_to_field")
                        remaining_count = sum(1 for i in all_issues if not i.get("archived"))

                        if converted_count > 0 or remaining_count > 0:
                            status_msg = json.dumps({
                                'step': total_steps - 1,
                                'message': f'📋 {converted_count} ایراد به فیلد تبدیل شد، {remaining_count} ایراد فعال باقی‌ماند',
                                'progress': 98
                            }, ensure_ascii=False)
                            yield f"data: {status_msg}\n\n"
                except Exception as arch_err:
                    import logging
                    logging.getLogger(__name__).error(f"Deep mode: Error checking issues: {arch_err}")

                # 🆕 ذخیره گزارش مهندسی 4 مرحله‌ای در دیتابیس
                report_id = f"eng_4step_{uuid.uuid4().hex[:12]}"
                try:
                    # جمع‌آوری نتایج مراحل (با استفاده از locals برای دسترسی امن)
                    local_vars = locals()
                    report_content = {
                        "depth": depth,
                        "models_used": selected_models,
                        "steps": {
                            "step1_validate_fields": local_vars.get('step1_result', {}),
                            "step2_health_to_fields": local_vars.get('step2_result', {}),
                            "step3_evaluate_models": local_vars.get('step3_result', {}),
                            "step4_update_roadmap": local_vars.get('step4_result', {}),
                        },
                        "completed_at": datetime.utcnow().isoformat(),
                    }

                    # ایجاد گزارش
                    report = Report(
                        id=report_id,
                        project_id=project_id,
                        report_type="engineering_4step",
                        title=f"گزارش مهندسی ۴ مرحله‌ای - {project.name if project else 'پروژه'}",
                        content=json.dumps(report_content, ensure_ascii=False, indent=2),
                        summary=f"گزارش مهندسی عمیق با {len(selected_models)} مدل",
                        total_activities=4,
                        total_tokens=0,  # در 4 مرحله‌ای توکن جداگانه حساب نشده
                        models_used=json.dumps(selected_models),
                        period_start=datetime.utcnow() - timedelta(days=days),
                        period_end=datetime.utcnow(),
                        created_at=datetime.utcnow(),
                        generated_by=",".join(selected_models),
                    )
                    db.add(report)

                    # ثبت در ژورنال
                    activity_log = ActivityLog(
                        id=f"log_{uuid.uuid4().hex[:12]}",
                        project_id=project_id,
                        model_id=selected_models[0] if selected_models else "unknown",
                        model_provider="multi_model",
                        activity_type="engineering_report_4step",
                        prompt=f"تولید گزارش مهندسی 4 مرحله‌ای برای {days} روز اخیر",
                        response=f"مراحل تکمیل شد: اعتبارسنجی فیلدها، تبدیل ایرادات، ارزیابی مدل‌ها، به‌روزرسانی نقشه راه",
                        tokens_used=0,
                        latency_ms=0,
                        success=True,
                        field_id=None,
                        field_name=f"گزارش مهندسی ۴ مرحله‌ای - {report_id}",
                        extra_data=json.dumps({
                            "report_id": report_id,
                            "depth": depth,
                            "models_used": selected_models,
                            "steps_completed": 4,
                        }, ensure_ascii=False),
                        created_at=datetime.utcnow()
                    )
                    db.add(activity_log)
                    db.commit()

                    logger.info(f"✅ 4-step engineering report saved: {report_id}")
                except Exception as save_err:
                    logger.error(f"Error saving 4-step report: {save_err}")

                result = {
                    "success": True,
                    "depth": depth,
                    "models_used": selected_models,
                    "steps_completed": 4,
                    "message": "گزارش مهندسی 4 مرحله‌ای تکمیل شد",
                    "report_id": report_id  # 🆕 شناسه گزارش
                }
            else:
                # حالت quick/standard: یک فراخوانی
                step = total_steps - 1
                models_text = ", ".join(selected_models)
                msg_final = json.dumps({'step': step, 'total': total_steps, 'message': f'🧠 تولید گزارش نهایی با {models_text}...', 'progress': 80}, ensure_ascii=False)
                yield f"data: {msg_final}\n\n"

                # فراخوانی گزارش اصلی (با مدل اول)
                result = await generate_engineering_report(
                    project_id=project_id,
                    days=days,
                    model_id=selected_models[0],
                    auto_create_fields=auto_create_fields,
                    validate_health_issues=validate_health_issues,
                    db=db
                )

                # 🆕 ثبت مدل‌های استفاده شده در نتیجه
                if result.get("success"):
                    result["models_used"] = selected_models
                    result["depth"] = depth

            # 🟢 اصلاح شده: فقط گزارش وضعیت ایرادات - بایگانی خودکار حذف شد
            # ایرادات فقط در صورت تبدیل به فیلد بایگانی می‌شوند (در generate_engineering_report)
            # بقیه ایرادات فعال می‌مانند تا کاربر بتواند آنها را دستی بررسی کند
            if result.get("success"):
                try:
                    project_fresh = db.query(Project).filter(Project.id == project_id).first()
                    if project_fresh and project_fresh.issues_found:
                        all_issues = json.loads(project_fresh.issues_found)
                        converted_count = sum(1 for i in all_issues if i.get("archived_reason") == "converted_to_field")
                        remaining_count = sum(1 for i in all_issues if not i.get("archived"))

                        result["issues_converted"] = converted_count
                        result["issues_remaining"] = remaining_count

                        if converted_count > 0 or remaining_count > 0:
                            status_msg = json.dumps({
                                'step': total_steps,
                                'message': f'📋 {converted_count} ایراد به فیلد تبدیل شد، {remaining_count} ایراد برای بررسی دستی باقی‌ماند',
                                'progress': 99
                            }, ensure_ascii=False)
                            yield f"data: {status_msg}\n\n"
                except Exception as stat_err:
                    import logging
                    logging.getLogger(__name__).error(f"Failed to get issues status: {stat_err}")

            # مرحله نهایی: اتمام
            if result.get("success"):
                success_msg = json.dumps({'step': total_steps, 'total': total_steps, 'message': '✅ گزارش با موفقیت تولید شد', 'progress': 100, 'result': result}, ensure_ascii=False)
                yield f"data: {success_msg}\n\n"
            else:
                error_text = result.get('error', 'خطای نامشخص')
                error_msg = json.dumps({'step': total_steps, 'message': f'❌ خطا: {error_text}', 'progress': 100, 'error': error_text}, ensure_ascii=False)
                yield f"data: {error_msg}\n\n"

        except Exception as e:
            exc_msg = json.dumps({'error': str(e), 'progress': 100}, ensure_ascii=False)
            yield f"data: {exc_msg}\n\n"

    return StreamingResponse(
        progress_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


# ===================== تریگر گزارش =====================

@router.get("/{project_id}/reports/trigger")
async def get_report_trigger(
    project_id: str,
    db: Session = Depends(get_db)
):
    """دریافت تنظیمات تریگر گزارش"""
    trigger = db.query(ReportTrigger).filter(ReportTrigger.project_id == project_id).first()

    if not trigger:
        return {
            "success": True,
            "trigger": {
                "enabled": False,
                "interval_minutes": 1440,
                "interval_type": "days",
                "report_model": "openai",
                "last_run": None,
                "next_run": None,
            }
        }

    return {
        "success": True,
        "trigger": {
            "enabled": trigger.enabled,
            "interval_minutes": trigger.interval_minutes,
            "interval_type": trigger.interval_type,
            "report_model": trigger.report_model,
            "last_run": trigger.last_run.isoformat() if trigger.last_run else None,
            "next_run": trigger.next_run.isoformat() if trigger.next_run else None,
        }
    }


@router.put("/{project_id}/reports/trigger")
async def update_report_trigger(
    project_id: str,
    settings: ReportTriggerSettings,
    db: Session = Depends(get_db)
):
    """بروزرسانی تنظیمات تریگر گزارش"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    trigger = db.query(ReportTrigger).filter(ReportTrigger.project_id == project_id).first()

    if not trigger:
        trigger = ReportTrigger(
            id=f"trigger_{uuid.uuid4().hex[:12]}",
            project_id=project_id,
        )
        db.add(trigger)

    trigger.enabled = settings.enabled
    trigger.interval_minutes = settings.interval_minutes
    trigger.interval_type = settings.interval_type
    trigger.report_model = settings.report_model

    # محاسبه زمان بعدی
    if settings.enabled:
        if settings.interval_type == "minutes":
            trigger.next_run = datetime.utcnow() + timedelta(minutes=settings.interval_minutes)
        elif settings.interval_type == "hours":
            trigger.next_run = datetime.utcnow() + timedelta(hours=settings.interval_minutes)
        else:
            trigger.next_run = datetime.utcnow() + timedelta(days=settings.interval_minutes)

    db.commit()

    return {
        "success": True,
        "message": "تنظیمات تریگر ذخیره شد",
        "trigger": {
            "enabled": trigger.enabled,
            "interval_minutes": trigger.interval_minutes,
            "interval_type": trigger.interval_type,
            "report_model": trigger.report_model,
            "next_run": trigger.next_run.isoformat() if trigger.next_run else None,
        }
    }


# ===================== 🆕 گزارش مهندسی ۴ مرحله‌ای =====================

@router.post("/{project_id}/engineering/step1-validate-fields")
async def engineering_step1_validate_fields(
    project_id: str,
    model_id: str = Query("claude"),
    depth: str = Query("normal", description="عمق بررسی: quick, normal, deep"),
    db: Session = Depends(get_db)
):
    """
    🔴 مرحله ۱ گزارش مهندسی: بررسی پروژه و اعتبارسنجی فیلدهای پویای موجود

    عملکرد:
    - کل پوشه اصلی پروژه بررسی می‌شود
    - فیلدهای پویای موجود ارزیابی می‌شوند
    - فیلدهای ضروری تاییدیه می‌گیرند
    - فیلدهای غیرضروری بایگانی می‌شوند
    - فیلدهای قابل ادغام ادغام می‌شوند
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.ai_base import Message
    import logging
    logger = logging.getLogger(__name__)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    # دریافت فیلدهای موجود
    existing_fields = []
    try:
        if project.dynamic_fields:
            existing_fields = json.loads(project.dynamic_fields)
    except:
        pass

    pending_fields = [f for f in existing_fields if not f.get("archived") and not f.get("engineering_approval")]
    approved_fields = [f for f in existing_fields if not f.get("archived") and f.get("engineering_approval")]

    # 🆕 ساخت prompt بر اساس عمق بررسی
    depth_instructions = {
        "quick": "بررسی سریع و خلاصه انجام بده. فقط موارد واضح را تایید یا رد کن.",
        "normal": "بررسی متعادل انجام بده. هر فیلد را بررسی کن اما وارد جزئیات نشو.",
        "deep": """بررسی بسیار عمیق و دقیق انجام بده:
- هر فیلد را خط به خط بررسی کن
- منطق و هدف هر فیلد را تحلیل کن
- تکراری بودن با فیلدهای موجود را دقیق بررسی کن
- امکان ادغام فیلدها را بررسی کن
- اولویت‌بندی را با دقت تعیین کن
- دلایل مفصل برای هر تصمیم بنویس
- کیفیت نوشتار و دستورات فیلدها را ارزیابی کن"""
    }

    system_prompt = f"""تو یک مهندس نرم‌افزار ارشد هستی. وظیفه‌ات بررسی و اعتبارسنجی فیلدهای پویای پروژه است.

🔴 سطح بررسی: {depth.upper()}
{depth_instructions.get(depth, depth_instructions["normal"])}

برای هر فیلد PENDING تصمیم بگیر:
1. approve: اگر فیلد ضروری و معتبر است
2. reject: اگر فیلد غیرضروری، تکراری یا نامعتبر است
3. merge: اگر با فیلد دیگری قابل ادغام است

خروجی JSON:
```json
{{
    "fields_to_approve": ["field_id1", "field_id2"],
    "fields_to_reject": [{{"id": "field_id", "reason": "دلیل رد"}}],
    "fields_to_merge": [{{"source_ids": ["id1", "id2"], "merged_name": "نام جدید", "merged_value": "دستور ادغام شده"}}],
    "fields_to_update": [{{"id": "field_id", "new_priority": 2, "new_action_type": "github_commit"}}],
    "analysis_details": {{"field_id": "تحلیل جزئی فیلد"}} if depth=="deep" else null,
    "summary": "خلاصه عملیات"
}}
```"""

    # 🆕 برای حالت deep، محتوای کامل‌تر فیلدها را نشان بده
    value_limit = 300 if depth != "deep" else 1000

    user_prompt = f"""پروژه: {project.name}
سطح بررسی: {depth.upper()}

=== فیلدهای PENDING (نیاز به تایید) ===
{json.dumps([{"id": f.get("id"), "name": f.get("name"), "value": f.get("value", "")[:value_limit], "action_type": f.get("action_type"), "target_path": f.get("target_path"), "priority": f.get("priority", 5)} for f in pending_fields], ensure_ascii=False, indent=2)}

=== فیلدهای تایید شده (برای مقایسه) ===
{json.dumps([{"id": f.get("id"), "name": f.get("name"), "action_type": f.get("action_type")} for f in approved_fields[:20]], ensure_ascii=False, indent=2)}

{"🔴 حالت DEEP: لطفاً هر فیلد را به صورت جداگانه و با جزئیات کامل تحلیل کن و دلیل تصمیمت را بنویس." if depth == "deep" else "لطفاً هر فیلد pending را بررسی و تصمیم بگیر."}"""

    ai_manager = get_ai_manager()

    # بررسی فعال بودن مدل
    if not ai_manager.get_enabled_status(model_id):
        fallback = ai_manager.find_fallback_model(model_id, task_type="engineering_step1")
        if fallback:
            model_id = fallback

    messages = [
        Message(role="system", content=system_prompt),
        Message(role="user", content=user_prompt),
    ]

    # 🆕 تنظیم max_tokens بر اساس عمق
    depth_tokens = {"quick": 2048, "normal": 4096, "deep": 8192}

    response = await ai_manager.generate(
        model_id=model_id,
        messages=messages,
        max_tokens=depth_tokens.get(depth, 4096),
        temperature=0.3 if depth != "deep" else 0.2,  # دقت بیشتر در deep
        task_type="engineering_step1",
        allow_fallback=True,
    )

    # پردازش نتایج
    result_data = {}
    try:
        import re
        content = response.content
        match = re.search(r'```(?:json)?\s*([\s\S]*?)```', content)
        if match:
            result_data = json.loads(match.group(1).strip())
        else:
            first = content.find('{')
            last = content.rfind('}')
            if first != -1 and last != -1:
                result_data = json.loads(content[first:last+1])
    except Exception as e:
        logger.error(f"Step1 JSON parse error: {e}")
        result_data = {"error": str(e)}

    # اعمال تغییرات
    approved_count = 0
    rejected_count = 0
    merged_count = 0

    for field_id in result_data.get("fields_to_approve", []):
        for field in existing_fields:
            if field.get("id") == field_id and not field.get("archived"):
                field["engineering_approval"] = {
                    "approved": True,
                    "approved_at": datetime.utcnow().isoformat(),
                    "approved_by": model_id,
                    "approval_type": "step1_validation",
                    "step": 1
                }
                field["validation_marker"] = "validated"
                approved_count += 1

    for reject_info in result_data.get("fields_to_reject", []):
        field_id = reject_info.get("id") if isinstance(reject_info, dict) else reject_info
        for field in existing_fields:
            if field.get("id") == field_id and not field.get("archived"):
                field["archived"] = True
                field["archived_at"] = datetime.utcnow().isoformat()
                field["archived_reason"] = reject_info.get("reason", "rejected_step1") if isinstance(reject_info, dict) else "rejected_step1"
                rejected_count += 1

    for merge_info in result_data.get("fields_to_merge", []):
        source_ids = merge_info.get("source_ids", [])
        if len(source_ids) >= 2:
            for sid in source_ids:
                for field in existing_fields:
                    if field.get("id") == sid:
                        field["archived"] = True
                        field["archived_at"] = datetime.utcnow().isoformat()
                        field["archived_reason"] = "merged_step1"

            merged_field = {
                "id": str(uuid.uuid4()),
                "name": merge_info.get("merged_name", "فیلد ادغام‌شده"),
                "value": merge_info.get("merged_value", ""),
                "target_models": ["claude"],
                "action_type": "github_commit",
                "field_type": "temporary",
                "priority": 3,
                "engineering_approval": {
                    "approved": True,
                    "approved_at": datetime.utcnow().isoformat(),
                    "approved_by": model_id,
                    "approval_type": "merged_step1",
                    "step": 1
                },
                "merged_from": source_ids,
            }
            existing_fields.append(merged_field)
            merged_count += 1

    # ذخیره تغییرات
    project.dynamic_fields = json.dumps(existing_fields, ensure_ascii=False)
    db.commit()

    # ثبت در ژورنال
    log_detailed_operation(
        db, project_id, None,
        "engineering_step1",
        f"مرحله ۱ گزارش مهندسی: {approved_count} تایید، {rejected_count} رد، {merged_count} ادغام",
        details={"approved": approved_count, "rejected": rejected_count, "merged": merged_count},
        status="completed"
    )
    db.commit()

    return {
        "success": True,
        "step": 1,
        "step_name": "validate_fields",
        "model_used": model_id,
        "results": {
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "merged_count": merged_count,
            "summary": result_data.get("summary", "")
        }
    }


@router.post("/{project_id}/engineering/step2-health-to-fields")
async def engineering_step2_health_to_fields(
    project_id: str,
    model_id: str = Query("claude"),
    depth: str = Query("normal", description="عمق بررسی: quick, normal, deep"),
    db: Session = Depends(get_db)
):
    """
    🔴 مرحله ۲ گزارش مهندسی: انطباق با تحلیل سلامت و تولید فیلدهای اقدام‌محور

    عملکرد:
    - ایرادات شناسایی‌شده در تحلیل سلامت تایید می‌شوند
    - امکان ادغام چند ایراد در یک فیلد بررسی می‌شود
    - فیلدهای جدید با تاییدیه گزارش مهندسی ایجاد می‌شوند
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.ai_base import Message
    import logging
    logger = logging.getLogger(__name__)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    # دریافت ایرادات سلامت - 🔴 نگه داشتن همه ایرادات (شامل بایگانی شده‌ها)
    all_health_issues = []
    if project.issues_found:
        try:
            all_health_issues = json.loads(project.issues_found)
        except:
            pass

    # فقط ایرادات بایگانی نشده برای پردازش
    active_issues = [i for i in all_health_issues if not i.get("archived")]

    if not active_issues:
        return {
            "success": True,
            "step": 2,
            "step_name": "health_to_fields",
            "message": "هیچ ایراد سلامت فعالی وجود ندارد",
            "results": {"fields_created": 0}
        }

    existing_fields = []
    try:
        if project.dynamic_fields:
            existing_fields = json.loads(project.dynamic_fields)
    except:
        pass

    # ساخت prompt
    system_prompt = """تو یک مهندس نرم‌افزار ارشد هستی. وظیفه‌ات اعتبارسنجی ایرادات سلامت و تبدیل آنها به فیلدهای عملیاتی است.

برای هر ایراد:
1. اگر معتبر است، create_field=true
2. اگر نامعتبر است، reject=true با دلیل
3. اگر چند ایراد قابل ادغام هستند، در یک فیلد ترکیب کن

خروجی JSON:
```json
{
    "validated_issues": [
        {
            "original_issue": {...},
            "validation_score": 85,
            "create_field": true,
            "field_details": {
                "name": "نام فیلد",
                "value": "دستور کامل برای AI",
                "action_type": "github_commit",
                "target_path": "path/to/file.py",
                "priority": 2
            }
        }
    ],
    "merged_issues": [
        {
            "source_issues": [{...}, {...}],
            "merged_field": {
                "name": "فیلد ادغام شده",
                "value": "دستور جامع",
                "action_type": "github_multi_commit",
                "priority": 2
            }
        }
    ],
    "rejected_issues": [
        {
            "original_issue": {...},
            "rejection_reason": "دلیل رد"
        }
    ],
    "summary": "خلاصه"
}
```

⚠️ فقط فیلدهای عملیاتی (github_commit/github_multi_commit) ایجاد کن، نه display!"""

    user_prompt = f"""پروژه: {project.name}

=== ایرادات سلامت ({len(active_issues)} عدد) ===
{json.dumps(active_issues[:50], ensure_ascii=False, indent=2)}

=== فیلدهای موجود (برای جلوگیری از تکرار) ===
{json.dumps([{"name": f.get("name"), "target_path": f.get("target_path")} for f in existing_fields if not f.get("archived")][:30], ensure_ascii=False, indent=2)}

ایرادات را بررسی و به فیلدهای عملیاتی تبدیل کن."""

    ai_manager = get_ai_manager()

    if not ai_manager.get_enabled_status(model_id):
        fallback = ai_manager.find_fallback_model(model_id, task_type="engineering_step2")
        if fallback:
            model_id = fallback

    messages = [
        Message(role="system", content=system_prompt),
        Message(role="user", content=user_prompt),
    ]

    response = await ai_manager.generate(
        model_id=model_id,
        messages=messages,
        max_tokens=8192,
        temperature=0.3,
        task_type="engineering_step2",
        allow_fallback=True,
    )

    # پردازش نتایج
    result_data = {}
    try:
        import re
        content = response.content
        match = re.search(r'```(?:json)?\s*([\s\S]*?)```', content)
        if match:
            result_data = json.loads(match.group(1).strip())
        else:
            first = content.find('{')
            last = content.rfind('}')
            if first != -1 and last != -1:
                result_data = json.loads(content[first:last+1])
    except Exception as e:
        logger.error(f"Step2 JSON parse error: {e}")
        result_data = {"error": str(e)}

    # ایجاد فیلدها از validated_issues
    created_fields = []
    for validated in result_data.get("validated_issues", []):
        if not validated.get("create_field"):
            continue

        field_details = validated.get("field_details", {})
        new_field = {
            "id": str(uuid.uuid4()),
            "name": field_details.get("name", "فیلد جدید"),
            "value": field_details.get("value", ""),
            "target_models": ["claude"],
            "action_type": field_details.get("action_type", "github_commit"),
            "target_path": field_details.get("target_path"),
            "field_type": "temporary",
            "priority": field_details.get("priority", 3),
            "archive_after_run": True,
            "engineering_approval": {
                "approved": True,
                "approved_at": datetime.utcnow().isoformat(),
                "approved_by": model_id,
                "approval_type": "health_validation_step2",
                "step": 2
            },
            "original_issue": validated.get("original_issue"),
            "validation_score": validated.get("validation_score", 0),
        }
        existing_fields.append(new_field)
        created_fields.append(new_field["name"])

        # بایگانی ایراد اصلی - 🆕 منطق تطبیق بهبود یافته
        orig = validated.get("original_issue", {})
        orig_file = orig.get("file") or orig.get("file_path") or ""
        orig_type = orig.get("type") or orig.get("severity") or ""
        orig_line = orig.get("line") or orig.get("start_line")
        orig_msg = (orig.get("message") or orig.get("description") or "")[:50]

        for issue in all_health_issues:
            issue_file = issue.get("file") or issue.get("file_path") or ""
            issue_type = issue.get("type") or issue.get("severity") or ""
            issue_line = issue.get("line") or issue.get("start_line")
            issue_msg = (issue.get("message") or issue.get("description") or "")[:50]

            # تطبیق انعطاف‌پذیر:
            # 1. فایل یکسان و پیام مشابه
            # 2. یا فایل یکسان و خط یکسان
            # 3. یا پیام کاملاً یکسان
            file_match = orig_file and issue_file and (orig_file in issue_file or issue_file in orig_file)
            line_match = orig_line and issue_line and orig_line == issue_line
            msg_match = orig_msg and issue_msg and (orig_msg in issue_msg or issue_msg in orig_msg)

            if (file_match and line_match) or (file_match and msg_match) or (issue.get("id") == orig.get("id")):
                issue["archived"] = True
                issue["archived_at"] = datetime.utcnow().isoformat()
                issue["archived_reason"] = "converted_to_field_step2"
                issue["converted_to_field"] = new_field["id"]

                # 🆕 ثبت در ژورنال برای هر ایراد بایگانی شده
                log_detailed_operation(
                    db, project_id, None,
                    "issue_archived",
                    f"ایراد تایید و تبدیل به فیلد شد: {issue.get('message', '')[:60]}",
                    details={
                        "issue_id": issue.get("id"),
                        "file": issue.get("file"),
                        "severity": issue.get("severity"),
                        "archive_reason": "approved_converted_to_field",
                        "created_field_id": new_field["id"],
                        "created_field_name": new_field["name"],
                        "validation_score": validated.get("validation_score", 0),
                    },
                    target_type="issue",
                    target_id=issue.get("id"),
                    target_name=issue.get("message", "")[:60],
                    status="completed"
                )
                break

    # ایجاد فیلدهای ادغام شده
    for merged in result_data.get("merged_issues", []):
        merged_field_details = merged.get("merged_field", {})
        new_field = {
            "id": str(uuid.uuid4()),
            "name": merged_field_details.get("name", "فیلد ادغام شده"),
            "value": merged_field_details.get("value", ""),
            "target_models": ["claude"],
            "action_type": merged_field_details.get("action_type", "github_multi_commit"),
            "field_type": "temporary",
            "priority": merged_field_details.get("priority", 2),
            "archive_after_run": True,
            "engineering_approval": {
                "approved": True,
                "approved_at": datetime.utcnow().isoformat(),
                "approved_by": model_id,
                "approval_type": "merged_health_step2",
                "step": 2
            },
            "source_issues": merged.get("source_issues"),
        }
        existing_fields.append(new_field)
        created_fields.append(new_field["name"])

    # بایگانی ایرادات رد شده - 🆕 منطق تطبیق بهبود یافته
    rejected_count = 0
    for rejected in result_data.get("rejected_issues", []):
        orig = rejected.get("original_issue", {})
        orig_file = orig.get("file") or orig.get("file_path") or ""
        orig_msg = (orig.get("message") or orig.get("description") or "")[:50]

        for issue in all_health_issues:
            issue_file = issue.get("file") or issue.get("file_path") or ""
            issue_msg = (issue.get("message") or issue.get("description") or "")[:50]

            # تطبیق انعطاف‌پذیر
            file_match = orig_file and issue_file and (orig_file in issue_file or issue_file in orig_file)
            msg_match = orig_msg and issue_msg and (orig_msg in issue_msg or issue_msg in orig_msg)

            if (file_match and msg_match) or (issue.get("id") == orig.get("id")):
                issue["archived"] = True
                issue["archived_at"] = datetime.utcnow().isoformat()
                rejection_reason = rejected.get('rejection_reason', '')
                issue["archived_reason"] = f"rejected_step2: {rejection_reason}"
                rejected_count += 1

                # 🆕 ثبت در ژورنال برای هر ایراد رد شده
                log_detailed_operation(
                    db, project_id, None,
                    "issue_rejected",
                    f"ایراد رد شد: {issue.get('message', '')[:60]}",
                    details={
                        "issue_id": issue.get("id"),
                        "file": issue.get("file"),
                        "severity": issue.get("severity"),
                        "archive_reason": "rejected",
                        "rejection_reason": rejection_reason,
                    },
                    target_type="issue",
                    target_id=issue.get("id"),
                    target_name=issue.get("message", "")[:60],
                    status="completed"
                )
                break

    # 🔴 FALLBACK: اگر هیچ ایرادی بایگانی نشده ولی فیلد ایجاد شده، همه active_issues را بایگانی کن
    archived_by_fallback = 0
    already_archived = sum(1 for i in all_health_issues if i.get("archived"))
    newly_archived = sum(1 for i in all_health_issues if i.get("archived_reason", "").startswith("converted_to_field_step2") or i.get("archived_reason", "").startswith("rejected_step2"))

    if len(created_fields) > 0 and newly_archived == 0:
        logger.warning(f"⚠️ FALLBACK: {len(created_fields)} fields created but no issues archived. Archiving all active issues.")
        for issue in all_health_issues:
            if not issue.get("archived"):
                issue["archived"] = True
                issue["archived_at"] = datetime.utcnow().isoformat()
                issue["archived_reason"] = "fallback_bulk_archive_step2"
                archived_by_fallback += 1

                # 🆕 ثبت در ژورنال برای هر ایراد بایگانی شده در fallback
                log_detailed_operation(
                    db, project_id, None,
                    "issue_archived",
                    f"ایراد بایگانی شد (fallback): {issue.get('message', '')[:60]}",
                    details={
                        "issue_id": issue.get("id"),
                        "file": issue.get("file"),
                        "severity": issue.get("severity"),
                        "archive_reason": "fallback_engineering_report",
                    },
                    target_type="issue",
                    target_id=issue.get("id"),
                    target_name=issue.get("message", "")[:60],
                    status="completed"
                )
        logger.info(f"✅ Fallback archived {archived_by_fallback} issues")

    total_archived = newly_archived + rejected_count + archived_by_fallback

    # ذخیره تغییرات
    project.dynamic_fields = json.dumps(existing_fields, ensure_ascii=False)
    project.issues_found = json.dumps(all_health_issues, ensure_ascii=False)
    db.commit()

    # ثبت در ژورنال
    log_detailed_operation(
        db, project_id, None,
        "engineering_step2",
        f"مرحله ۲: {len(created_fields)} فیلد ایجاد، {total_archived} ایراد بایگانی شد",
        details={"created": len(created_fields), "rejected": rejected_count, "archived": total_archived, "fallback": archived_by_fallback},
        status="completed"
    )
    db.commit()

    return {
        "success": True,
        "step": 2,
        "step_name": "health_to_fields",
        "model_used": model_id,
        "created_count": len(created_fields),
        "archived_count": total_archived,
        "results": {
            "fields_created": len(created_fields),
            "fields_created_names": created_fields,
            "rejected_count": rejected_count,
            "archived_count": total_archived,
            "fallback_archived": archived_by_fallback,
            "summary": result_data.get("summary", "")
        }
    }


@router.post("/{project_id}/engineering/step3-evaluate-models")
async def engineering_step3_evaluate_models(
    project_id: str,
    model_id: str = Query("claude"),
    depth: str = Query("normal", description="عمق بررسی: quick, normal, deep"),
    db: Session = Depends(get_db)
):
    """
    🔴 مرحله ۳ گزارش مهندسی: اعتبارسنجی عملکرد مدل‌های تحلیل سلامت و به‌روزرسانی ساختار

    عملکرد:
    - عملکرد مدل‌های تحلیل سلامت ارزیابی می‌شود
    - امتیاز مثبت/منفی به مدل‌ها داده می‌شود
    - ساختار و رنگ‌بندی به‌روز می‌شود
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.model_profiler import ModelProfiler
    import logging
    logger = logging.getLogger(__name__)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    profiler = ModelProfiler()

    # دریافت نتایج اعتبارسنجی قبلی
    validation_results = {}
    if project.last_validation_results:
        try:
            validation_results = json.loads(project.last_validation_results)
        except:
            pass

    # دریافت ایرادات برای شمارش
    health_issues = []
    if project.issues_found:
        try:
            health_issues = json.loads(project.issues_found)
        except:
            pass

    # شمارش ایرادات هر مدل
    model_stats = {}
    for issue in health_issues:
        source_models = issue.get("source_models", [issue.get("source_model", "unknown")])
        if isinstance(source_models, str):
            source_models = [source_models]

        is_valid = not issue.get("archived") or issue.get("archived_reason", "").startswith("converted")
        is_rejected = issue.get("archived") and "rejected" in issue.get("archived_reason", "")

        for model in source_models:
            if model not in model_stats:
                model_stats[model] = {"correct": 0, "false_positive": 0, "total": 0}
            model_stats[model]["total"] += 1
            if is_valid:
                model_stats[model]["correct"] += 1
            elif is_rejected:
                model_stats[model]["false_positive"] += 1

    # به‌روزرسانی پروفایل مدل‌ها
    updated_profiles = []
    for model_id_stat, stats in model_stats.items():
        if model_id_stat == "unknown" or stats["total"] == 0:
            continue

        try:
            # محاسبه precision قبل از به‌روزرسانی
            precision = round(stats["correct"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0
            false_positive_rate = round(stats["false_positive"] / stats["total"] * 100, 1) if stats["total"] > 0 else 0

            # تعیین تغییر امتیاز (مثبت یا منفی)
            if stats["false_positive"] > stats["correct"]:
                score_change = "منفی"
                score_reason = f"تعداد ایرادات نادرست ({stats['false_positive']}) بیشتر از صحیح ({stats['correct']})"
            elif precision >= 80:
                score_change = "مثبت"
                score_reason = f"دقت بالا ({precision}%) - عملکرد عالی"
            elif precision >= 60:
                score_change = "خنثی"
                score_reason = f"دقت متوسط ({precision}%) - قابل بهبود"
            else:
                score_change = "منفی"
                score_reason = f"دقت پایین ({precision}%) - نیاز به بازبینی"

            await profiler.update_profile(
                model_id=model_id_stat,
                task_type="health_analysis",
                correct_findings=stats["correct"],
                total_expected=stats["total"],
                false_positives=stats["false_positive"],
                response_time=0,
                tokens_used=0,
                details={
                    "project_id": project_id,
                    "evaluation_step": 3,
                    "evaluation_date": datetime.utcnow().isoformat(),
                }
            )
            updated_profiles.append({
                "model_id": model_id_stat,
                "correct": stats["correct"],
                "false_positive": stats["false_positive"],
                "precision": precision
            })
            logger.info(f"Updated profile for {model_id_stat}: {stats}")

            # 🆕 ثبت در ژورنال برای هر تغییر امتیاز مدل
            log_detailed_operation(
                db, project_id, None,
                "model_score_updated",
                f"امتیاز مدل {model_id_stat}: {score_change} - دقت {precision}%",
                details={
                    "model_id": model_id_stat,
                    "score_change_type": score_change,
                    "precision": precision,
                    "false_positive_rate": false_positive_rate,
                    "correct_findings": stats["correct"],
                    "false_positives": stats["false_positive"],
                    "total_issues": stats["total"],
                    "reason": score_reason,
                    "evaluation_step": 3,
                },
                target_type="model",
                target_id=model_id_stat,
                target_name=model_id_stat,
                status="completed"
            )
        except Exception as e:
            logger.warning(f"Could not update profile for {model_id_stat}: {e}")

    # به‌روزرسانی file_health_map
    file_health_map = {}
    if project.file_health_map:
        try:
            file_health_map = json.loads(project.file_health_map)
        except:
            pass

    # شمارش ایرادات فعال هر فایل
    for issue in health_issues:
        if issue.get("archived"):
            continue
        file_path = issue.get("file", "")
        if file_path not in file_health_map:
            file_health_map[file_path] = {"score": 100, "issues": 0}

        file_health_map[file_path]["issues"] = file_health_map[file_path].get("issues", 0) + 1
        # کاهش امتیاز براساس تعداد ایرادات
        severity = issue.get("severity", "medium")
        penalty = {"critical": 25, "high": 15, "medium": 10, "low": 5}.get(severity, 10)
        file_health_map[file_path]["score"] = max(0, file_health_map[file_path].get("score", 100) - penalty)

    # محاسبه رنگ براساس امتیاز
    for file_path, data in file_health_map.items():
        score = data.get("score", 100)
        if score >= 80:
            data["color"] = "#22c55e"  # سبز
            data["hex"] = "#22c55e"
        elif score >= 60:
            data["color"] = "#eab308"  # زرد
            data["hex"] = "#eab308"
        elif score >= 40:
            data["color"] = "#f97316"  # نارنجی
            data["hex"] = "#f97316"
        else:
            data["color"] = "#ef4444"  # قرمز
            data["hex"] = "#ef4444"

    project.file_health_map = json.dumps(file_health_map, ensure_ascii=False)
    db.commit()

    # ثبت در ژورنال
    log_detailed_operation(
        db, project_id, None,
        "engineering_step3",
        f"مرحله ۳: ارزیابی {len(updated_profiles)} مدل، به‌روزرسانی رنگ {len(file_health_map)} فایل",
        details={"models_evaluated": len(updated_profiles), "files_updated": len(file_health_map)},
        status="completed"
    )
    db.commit()

    return {
        "success": True,
        "step": 3,
        "step_name": "evaluate_models",
        "results": {
            "models_evaluated": updated_profiles,
            "files_color_updated": len(file_health_map),
            "file_health_sample": dict(list(file_health_map.items())[:10])
        }
    }


@router.post("/{project_id}/engineering/step4-update-roadmap")
async def engineering_step4_update_roadmap(
    project_id: str,
    model_id: str = Query("claude"),
    db: Session = Depends(get_db)
):
    """
    🔴 مرحله ۴ گزارش مهندسی: به‌روزرسانی نقشه راه و تعیین حالت ایده‌آل

    عملکرد:
    - نقشه راه به‌روز می‌شود با چک‌باکس‌ها
    - حالت ایده‌آل پروژه تعیین می‌شود
    - فیلدهای عملیاتی برای آیتم‌های ناقص ایجاد می‌شود
    """
    from ...services.ai_manager import get_ai_manager
    from ...services.ai_base import Message
    import logging
    logger = logging.getLogger(__name__)

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    existing_fields = []
    try:
        if project.dynamic_fields:
            existing_fields = json.loads(project.dynamic_fields)
    except:
        pass

    # فیلدهای اجرا شده
    executed_fields = [f for f in existing_fields if f.get("executed") or (f.get("archived") and "executed" in f.get("archived_reason", ""))]

    # ساخت prompt
    system_prompt = """تو یک مهندس نرم‌افزار ارشد هستی. وظیفه‌ات به‌روزرسانی نقشه راه و تعیین حالت ایده‌آل است.

خروجی JSON:
```json
{
    "roadmap_updates": [
        {"item": "نام آیتم", "completed": true, "reason": "فیلد X اجرا شد"},
        {"item": "آیتم ناقص", "completed": false, "create_field": true, "field_details": {"name": "...", "value": "...", "action_type": "github_commit", "target_path": "...", "priority": 2}}
    ],
    "ideal_state": {
        "description": "توضیح حالت ایده‌آل (3-5 پاراگراف)",
        "current_deficiencies": ["کمبود 1", "کمبود 2"],
        "required_actions": ["اقدام 1", "اقدام 2"],
        "target_architecture": "توضیح ساختار هدف"
    },
    "new_roadmap_content": "محتوای کامل نقشه راه به‌روز شده با چک‌باکس‌ها",
    "summary": "خلاصه"
}
```

آیتم‌های تکمیل شده: [x] ✅
آیتم‌های ناقص: [ ]"""

    user_prompt = f"""پروژه: {project.name}
نوع: {project.project_type or 'نامشخص'}

=== نقشه راه فعلی ===
{project.roadmap_content or 'تعریف نشده'}

=== فیلدهای اجرا شده ===
{json.dumps([{"name": f.get("name"), "target_path": f.get("target_path")} for f in executed_fields[:30]], ensure_ascii=False, indent=2)}

=== حالت ایده‌آل فعلی ===
{project.ideal_state or 'تعریف نشده'}

نقشه راه را به‌روز کن و حالت ایده‌آل جدید را تعیین کن."""

    ai_manager = get_ai_manager()

    if not ai_manager.get_enabled_status(model_id):
        fallback = ai_manager.find_fallback_model(model_id, task_type="engineering_step4")
        if fallback:
            model_id = fallback

    messages = [
        Message(role="system", content=system_prompt),
        Message(role="user", content=user_prompt),
    ]

    response = await ai_manager.generate(
        model_id=model_id,
        messages=messages,
        max_tokens=8192,
        temperature=0.3,
        task_type="engineering_step4",
        allow_fallback=True,
    )

    # پردازش نتایج
    result_data = {}
    try:
        import re
        content = response.content
        match = re.search(r'```(?:json)?\s*([\s\S]*?)```', content)
        if match:
            result_data = json.loads(match.group(1).strip())
        else:
            first = content.find('{')
            last = content.rfind('}')
            if first != -1 and last != -1:
                result_data = json.loads(content[first:last+1])
    except Exception as e:
        logger.error(f"Step4 JSON parse error: {e}")
        result_data = {"error": str(e)}

    # به‌روزرسانی نقشه راه
    new_roadmap = result_data.get("new_roadmap_content")
    if new_roadmap:
        project.roadmap_content = new_roadmap
    else:
        # 🆕 اگر AI محتوای نقشه راه را برنگرداند، خودمان از updates بسازیم
        roadmap_updates = result_data.get("roadmap_updates", [])
        if roadmap_updates:
            roadmap_lines = [f"# نقشه راه پروژه {project.name}\n"]
            roadmap_lines.append(f"*به‌روزرسانی: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}*\n")
            for update in roadmap_updates:
                status = "[x] ✅" if update.get("completed") else "[ ]"
                roadmap_lines.append(f"{status} {update.get('item', 'آیتم')}")
                if update.get("reason"):
                    roadmap_lines.append(f"   - {update.get('reason')}")
            project.roadmap_content = "\n".join(roadmap_lines)
            new_roadmap = project.roadmap_content
            logger.info(f"Generated roadmap from updates: {len(roadmap_updates)} items")
        else:
            # 🆕 Fallback نهایی: ساخت نقشه راه از وضعیت فعلی پروژه
            roadmap_lines = [f"# نقشه راه پروژه {project.name}\n"]
            roadmap_lines.append(f"*تولید خودکار: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}*\n")
            roadmap_lines.append("\n## وضعیت فعلی\n")

            # فیلدهای تایید شده
            approved_count = len([f for f in existing_fields if f.get("engineering_approval", {}).get("approved")])
            executed_count = len(executed_fields)
            pending_count = len([f for f in existing_fields if not f.get("archived") and not f.get("engineering_approval")])

            if approved_count > 0:
                roadmap_lines.append(f"[x] ✅ {approved_count} فیلد تایید شده توسط گزارش مهندسی")
            if executed_count > 0:
                roadmap_lines.append(f"[x] ✅ {executed_count} فیلد اجرا شده")
            if pending_count > 0:
                roadmap_lines.append(f"[ ] {pending_count} فیلد در انتظار بررسی")

            # ایرادات
            health_issues = []
            try:
                if project.issues_found:
                    health_issues = json.loads(project.issues_found)
            except:
                pass
            active_issues = [i for i in health_issues if not i.get("archived")]
            if active_issues:
                roadmap_lines.append(f"[ ] {len(active_issues)} ایراد سلامت نیاز به رسیدگی")

            roadmap_lines.append("\n## مراحل بعدی\n")
            roadmap_lines.append("[ ] اجرای فیلدهای تایید شده")
            roadmap_lines.append("[ ] رفع ایرادات سلامت شناسایی شده")
            roadmap_lines.append("[ ] اجرای مجدد گزارش مهندسی پس از تغییرات")

            project.roadmap_content = "\n".join(roadmap_lines)
            new_roadmap = project.roadmap_content
            logger.info(f"Generated fallback roadmap from project state")

    # به‌روزرسانی حالت ایده‌آل
    ideal_state = result_data.get("ideal_state", {})
    if ideal_state:
        ideal_text = f"""## حالت ایده‌آل پروژه

{ideal_state.get('description', '')}

### کمبودهای فعلی:
{chr(10).join(['- ' + d for d in ideal_state.get('current_deficiencies', [])])}

### اقدامات مورد نیاز:
{chr(10).join(['- ' + a for a in ideal_state.get('required_actions', [])])}

### ساختار هدف:
{ideal_state.get('target_architecture', '')}
"""
        project.ideal_state = ideal_text

    # ایجاد فیلد برای آیتم‌های ناقص
    created_fields = []
    for update in result_data.get("roadmap_updates", []):
        if not update.get("completed") and update.get("create_field"):
            field_details = update.get("field_details", {})
            new_field = {
                "id": str(uuid.uuid4()),
                "name": f"[نقشه راه] {field_details.get('name', update.get('item', 'تسک'))}",
                "value": field_details.get("value", f"تسک نقشه راه: {update.get('item', '')}"),
                "target_models": ["claude"],
                "action_type": field_details.get("action_type", "github_commit"),
                "target_path": field_details.get("target_path"),
                "field_type": "temporary",
                "priority": field_details.get("priority", 3),
                "archive_after_run": True,
                "engineering_approval": {
                    "approved": True,
                    "approved_at": datetime.utcnow().isoformat(),
                    "approved_by": model_id,
                    "approval_type": "roadmap_step4",
                    "step": 4
                },
                "source": "roadmap",
                "roadmap_item": update.get("item"),
            }
            existing_fields.append(new_field)
            created_fields.append(new_field["name"])

    project.dynamic_fields = json.dumps(existing_fields, ensure_ascii=False)
    db.commit()

    # ثبت در ژورنال
    log_detailed_operation(
        db, project_id, None,
        "engineering_step4",
        f"مرحله ۴: نقشه راه به‌روز شد، {len(created_fields)} فیلد از نقشه راه ایجاد شد",
        details={"fields_created": len(created_fields), "roadmap_updated": bool(new_roadmap), "ideal_state_updated": bool(ideal_state)},
        status="completed"
    )
    db.commit()

    return {
        "success": True,
        "step": 4,
        "step_name": "update_roadmap",
        "model_used": model_id,
        "results": {
            "roadmap_updated": bool(new_roadmap),
            "ideal_state_updated": bool(ideal_state),
            "fields_created": len(created_fields),
            "fields_created_names": created_fields,
            "summary": result_data.get("summary", "")
        }
    }


class MultiModelEngineeringRequest(BaseModel):
    """درخواست گزارش مهندسی با چند مدل"""
    model_ids: List[str] = ["claude"]  # لیست مدل‌ها
    parallel: bool = False  # اجرای موازی یا ترتیبی


@router.post("/{project_id}/engineering/run-all-steps")
async def engineering_run_all_steps(
    project_id: str,
    model_id: str = Query("claude"),
    model_ids: str = Query(None, description="لیست مدل‌ها با کاما جدا شده (مثال: claude,gpt-4,gemini)"),
    depth: str = Query("normal", description="عمق تحلیل: quick, normal, deep"),
    db: Session = Depends(get_db)
):
    """
    🔴 اجرای تمام ۴ مرحله گزارش مهندسی به ترتیب

    🆕 قابلیت‌های جدید:
    - پشتیبانی از چند مدل همزمان (model_ids با کاما جدا شوند)
    - جایگزینی خودکار مدل‌های غیرفعال
    - ثبت جایگزینی در ژورنال
    - 🆕 پارامتر depth برای تعیین عمق تحلیل:
      - quick: بررسی سریع (فقط مراحل 1 و 4)
      - normal: بررسی معمول (تمام مراحل)
      - deep: بررسی عمیق (تمام مراحل + چندین تکرار)
    """
    results = {}
    replacement_notes = []

    # 🆕 پردازش لیست مدل‌ها
    selected_models = []
    if model_ids:
        selected_models = [m.strip() for m in model_ids.split(",") if m.strip()]
    else:
        selected_models = [model_id]

    # 🆕 اطمینان از فعال بودن مدل‌ها و جایگزینی در صورت نیاز
    active_selected = []
    for m in selected_models:
        final_model, was_replaced, note = ensure_active_model(db, m, "engineering_report")
        active_selected.append(final_model)
        if was_replaced:
            replacement_notes.append(note)
            # ثبت جایگزینی در ژورنال
            log_detailed_operation(
                db, project_id, None,
                "model_replacement",
                note,
                details={"original": m, "replacement": final_model, "reason": "model_disabled"},
                status="completed"
            )

    # حذف تکراری‌ها
    active_selected = list(dict.fromkeys(active_selected))

    # انتخاب مدل اصلی (اولین مدل فعال)
    primary_model = active_selected[0] if active_selected else "claude"

    results["models_used"] = active_selected
    results["replacement_notes"] = replacement_notes
    results["depth"] = depth

    # 🆕 تنظیمات بر اساس عمق تحلیل
    import asyncio
    depth_settings = {
        "quick": {"iterations": 1, "skip_steps": [2, 3], "max_tokens": 2048, "pause_seconds": 1},
        "normal": {"iterations": 1, "skip_steps": [], "max_tokens": 4096, "pause_seconds": 2},
        "deep": {"iterations": 3, "skip_steps": [], "max_tokens": 8192, "pause_seconds": 8},  # 🆕 3 تکرار با توقف 8 ثانیه
    }
    settings = depth_settings.get(depth, depth_settings["normal"])
    pause_seconds = settings.get("pause_seconds", 2)

    # 🆕 اگر چند مدل انتخاب شده، از هر کدام برای یک مرحله استفاده کن
    # یا از همه برای اعتبارسنجی متقابل
    model_for_step = {
        1: active_selected[0] if len(active_selected) > 0 else primary_model,
        2: active_selected[1 % len(active_selected)] if len(active_selected) > 1 else primary_model,
        3: primary_model,  # مرحله 3 ارزیابی همه مدل‌هاست
        4: active_selected[-1] if active_selected else primary_model,
    }

    # 🆕 در حالت deep، تحلیل چندمرحله‌ای واقعی انجام می‌شود
    total_steps_done = 0

    # 🆕 اجرای مراحل بر اساس عمق تحلیل
    for iteration in range(settings["iterations"]):
        iteration_results = {}

        if settings["iterations"] > 1:
            log_detailed_operation(
                db, project_id, None,
                "engineering_iteration",
                f"شروع تکرار {iteration + 1} از {settings['iterations']} (حالت: {depth})",
                status="in_progress"
            )
            db.commit()
            # 🆕 توقف واقعی بین تکرارها
            await asyncio.sleep(pause_seconds)

        # مرحله ۱
        if 1 not in settings["skip_steps"]:
            try:
                log_detailed_operation(db, project_id, None, "step1_progress", f"شروع مرحله ۱ - تکرار {iteration+1}", status="in_progress")
                db.commit()
                step1 = await engineering_step1_validate_fields(project_id, model_for_step[1], depth, db)
                iteration_results["step1"] = step1
                results["step1"] = step1
                total_steps_done += 1
                # 🆕 توقف بعد از هر مرحله در حالت deep
                if depth == "deep":
                    await asyncio.sleep(pause_seconds // 2)
            except Exception as e:
                results["step1"] = {"success": False, "error": str(e)}

        # مرحله ۲
        if 2 not in settings["skip_steps"]:
            try:
                log_detailed_operation(db, project_id, None, "step2_progress", f"شروع مرحله ۲ - تکرار {iteration+1}", status="in_progress")
                db.commit()
                step2 = await engineering_step2_health_to_fields(project_id, model_for_step[2], depth, db)
                iteration_results["step2"] = step2
                results["step2"] = step2
                total_steps_done += 1
                # 🆕 توقف بعد از هر مرحله در حالت deep
                if depth == "deep":
                    await asyncio.sleep(pause_seconds // 2)
            except Exception as e:
                results["step2"] = {"success": False, "error": str(e)}
        else:
            results["step2"] = {"success": True, "skipped": True, "message": "رد شده در حالت quick"}

        # مرحله ۳ - 🆕 ارزیابی با امتیازدهی دقیق
        if 3 not in settings["skip_steps"]:
            try:
                log_detailed_operation(db, project_id, None, "step3_progress", f"شروع مرحله ۳ - تکرار {iteration+1}", status="in_progress")
                db.commit()
                step3 = await engineering_step3_evaluate_models(project_id, model_for_step[3], depth, db)
                iteration_results["step3"] = step3
                results["step3"] = step3
                total_steps_done += 1
                # 🆕 توقف بعد از هر مرحله در حالت deep
                if depth == "deep":
                    await asyncio.sleep(pause_seconds // 2)
            except Exception as e:
                results["step3"] = {"success": False, "error": str(e)}
        else:
            results["step3"] = {"success": True, "skipped": True, "message": "رد شده در حالت quick"}

        # 🆕 در حالت deep، بین تکرارها توقف طولانی‌تر کن
        if iteration < settings["iterations"] - 1:
            log_detailed_operation(
                db, project_id, None,
                "engineering_pause",
                f"تکرار {iteration + 1} تکمیل شد - آماده‌سازی تکرار بعدی...",
                status="completed"
            )
            db.commit()
            await asyncio.sleep(pause_seconds)

    # مرحله ۴ (همیشه اجرا می‌شود) - 🆕 با توقف قبل از آن در حالت deep
    if depth == "deep":
        log_detailed_operation(db, project_id, None, "step4_prep", "آماده‌سازی مرحله نهایی نقشه راه...", status="in_progress")
        db.commit()
        await asyncio.sleep(pause_seconds // 2)

    try:
        step4 = await engineering_step4_update_roadmap(project_id, model_for_step[4], db)
        results["step4"] = step4
        total_steps_done += 1
    except Exception as e:
        results["step4"] = {"success": False, "error": str(e)}

    results["total_steps_executed"] = total_steps_done
    results["iterations_completed"] = settings["iterations"]

    all_success = all(results.get(f"step{i}", {}).get("success", False) for i in range(1, 5))

    db.commit()

    return {
        "success": all_success,
        "message": "تمام ۴ مرحله گزارش مهندسی اجرا شد" if all_success else "برخی مراحل با خطا مواجه شدند",
        "models_used": active_selected,
        "replacement_notes": replacement_notes if replacement_notes else None,
        "steps": results
    }


# ===================== جزئیات گزارش (باید آخر باشد) =====================

@router.get("/{project_id}/reports/{report_id}")
async def get_report_detail(
    project_id: str,
    report_id: str,
    db: Session = Depends(get_db)
):
    """دریافت جزئیات کامل یک گزارش"""
    report = db.query(Report).filter(
        Report.id == report_id,
        Report.project_id == project_id
    ).first()

    if not report:
        raise HTTPException(status_code=404, detail="گزارش یافت نشد")

    models_used = []
    try:
        if report.models_used:
            models_used = json.loads(report.models_used)
    except:
        pass

    return {
        "success": True,
        "report": {
            "id": report.id,
            "report_type": report.report_type,
            "title": report.title,
            "content": report.content,
            "summary": report.summary,
            "total_activities": report.total_activities,
            "total_tokens": report.total_tokens,
            "models_used": models_used,
            "period_start": report.period_start.isoformat() if report.period_start else None,
            "period_end": report.period_end.isoformat() if report.period_end else None,
            "created_at": report.created_at.isoformat() if report.created_at else None,
            "generated_by": report.generated_by,
        }
    }


# =====================================
# 🔴 نقشه راه - منتقل شده از تب تحلیل سلامت
# =====================================

class RoadmapRequest(BaseModel):
    """درخواست به‌روزرسانی نقشه راه"""
    content: Optional[str] = None
    auto_generate: bool = False
    model_id: Optional[str] = None


@router.get("/{project_id}/roadmap")
async def get_project_roadmap(project_id: str, db=Depends(get_db)):
    """
    دریافت نقشه راه پروژه
    🔴 این endpoint از تب تحلیل سلامت به تب ژورنال و گزارشات منتقل شده
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    issues = []
    try:
        if project.issues_found:
            issues = json.loads(project.issues_found)
    except:
        pass

    # دریافت آخرین گزارش مهندسی برای اطلاعات تکمیلی
    last_report = db.query(Report).filter(
        Report.project_id == project_id,
        Report.report_type == "engineering"
    ).order_by(desc(Report.created_at)).first()

    return {
        "success": True,
        "project_id": project_id,
        "roadmap_exists": bool(project.roadmap_content),
        "roadmap_content": project.roadmap_content or "",
        "ideal_state": project.ideal_state or "",
        "issues_found": issues,
        "last_engineering_report": {
            "id": last_report.id,
            "created_at": last_report.created_at.isoformat() if last_report and last_report.created_at else None,
            "summary": last_report.summary if last_report else None
        } if last_report else None
    }


@router.put("/{project_id}/roadmap")
async def update_project_roadmap(
    project_id: str,
    request: RoadmapRequest,
    db=Depends(get_db)
):
    """
    به‌روزرسانی نقشه راه پروژه
    🔴 تحلیل سلامت نقشه راه را به‌روزرسانی نمی‌کند - فقط گزارش مهندسی
    """
    from ...models.project import ProjectFile

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    if request.content:
        # ذخیره دستی
        project.roadmap_content = request.content
        db.commit()

        # ثبت در ژورنال
        log_entry = ActivityLog(
            id=f"log_{uuid.uuid4().hex[:12]}",
            project_id=project_id,
            model_id="manual",
            activity_type="roadmap_update",
            prompt="به‌روزرسانی دستی نقشه راه",
            response=f"نقشه راه ذخیره شد ({len(request.content)} کاراکتر)",
            tokens_used=0,
            latency_ms=0,
            success=True,
            created_at=datetime.utcnow(),
        )
        db.add(log_entry)
        db.commit()

        return {
            "success": True,
            "message": "نقشه راه ذخیره شد",
            "source": "manual"
        }

    elif request.auto_generate:
        # تولید خودکار با AI
        from ...services.ai_manager import get_ai_manager
        from ...services.ai_base import Message

        ai_manager = get_ai_manager()
        model_id = request.model_id or "claude"

        files = db.query(ProjectFile).filter(ProjectFile.project_id == project_id).all()
        files_list = [f.file_path for f in files[:50]]

        # پرامپت برای تولید نقشه راه
        prompt = f"""پروژه: {project.name}
توضیحات: {project.description or 'ندارد'}
نوع: {project.project_type or 'نامشخص'}

فایل‌های پروژه:
{chr(10).join(['- ' + f for f in files_list])}

حالت ایده‌آل فعلی:
{project.ideal_state or 'تعریف نشده'}

یک نقشه راه کامل و حرفه‌ای به زبان فارسی بنویس که شامل:
1. اهداف کوتاه‌مدت و بلندمدت
2. فازهای توسعه با جزئیات
3. موارد انجام شده (با تیک سبز ✅)
4. موارد در انتظار (با باکس خالی ⬜)
5. زمان‌بندی تقریبی

فرمت: Markdown با چک‌باکس‌ها"""

        try:
            response = await ai_manager.generate(
                model_id=model_id,
                messages=[Message(role="user", content=prompt)],
                max_tokens=3000,
                temperature=0.5
            )

            content = response.content.strip()
            if content.startswith("```"):
                content = content.split("```", 2)[1]
                if content.startswith("markdown"):
                    content = content[8:]
                content = content.strip()

            project.roadmap_content = content
            db.commit()

            # ثبت در ژورنال
            log_entry = ActivityLog(
                id=f"log_{uuid.uuid4().hex[:12]}",
                project_id=project_id,
                model_id=model_id,
                activity_type="roadmap_generation",
                prompt=prompt[:500],
                response=content[:1000],
                tokens_used=response.tokens_used or 0,
                latency_ms=0,
                success=True,
                created_at=datetime.utcnow(),
            )
            db.add(log_entry)
            db.commit()

            return {
                "success": True,
                "message": "نقشه راه تولید شد",
                "source": "auto",
                "model_used": model_id
            }

        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

    return {"success": False, "message": "محتوا یا auto_generate لازم است"}


# =====================================
# 🔴 تابع کمکی برای ثبت در ژورنال با پشتیبانی از fallback
# =====================================

def log_activity_with_fallback(
    db,
    project_id: str,
    model_id: str,
    activity_type: str,
    prompt: str,
    response_content: str,
    tokens_used: int = 0,
    latency_ms: int = 0,
    success: bool = True,
    error_message: str = None,
    field_id: str = None,
    field_name: str = None,
    fallback_used: bool = False,
    original_model_id: str = None,
    extra_data: dict = None
):
    """
    ثبت فعالیت در ژورنال با پشتیبانی از اطلاعات fallback

    🔴 اگر مدلی غیرفعال شده و fallback استفاده شده:
    - نام مدل جایگزین ثبت می‌شود
    - توضیح جایگزینی در extra_data ذخیره می‌شود
    """
    extra = extra_data or {}

    # 🔴 ثبت اطلاعات fallback
    if fallback_used and original_model_id:
        extra["fallback_info"] = {
            "original_model": original_model_id,
            "replacement_model": model_id,
            "reason": "مدل اصلی غیرفعال بود - جایگزین خودکار",
            "timestamp": datetime.utcnow().isoformat()
        }
        # 🔴 اصلاح نام مدل در provider برای نمایش صحیح
        provider = model_id.split("-")[0] if "-" in model_id else model_id
    else:
        provider = model_id.split("-")[0] if "-" in model_id else model_id

    log_entry = ActivityLog(
        id=f"log_{uuid.uuid4().hex[:12]}",
        project_id=project_id,
        model_id=model_id,
        model_provider=provider,
        activity_type=activity_type,
        prompt=prompt[:2000] if prompt else None,
        response=response_content[:5000] if response_content else None,
        tokens_used=tokens_used,
        latency_ms=latency_ms,
        success=success,
        error_message=error_message[:500] if error_message else None,
        field_id=field_id,
        field_name=field_name,
        created_at=datetime.utcnow(),
        extra_data=json.dumps(extra, ensure_ascii=False) if extra else None
    )

    db.add(log_entry)
    return log_entry


# =====================================
# 🆕 جایگزینی خودکار مدل غیرفعال در ژورنال
# =====================================

@router.post("/{project_id}/journal/fix-disabled-models")
async def fix_disabled_models_in_journal(
    project_id: str,
    db: Session = Depends(get_db)
):
    """
    🆕 بررسی و اصلاح مدل‌های غیرفعال در ژورنال

    اگر مدلی در صفحه مدل‌ها غیرفعال شده اما در ژورنال فعالیت آن مشاهده می‌شود:
    - سیستم به طور خودکار مدل جایگزین فعال را برای آن عملیات اختصاص می‌دهد
    - نام مدل قبلی در extra_data ثبت می‌شود
    - توضیح جایگزینی در ژورنال ذخیره می‌شود
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    from ...services.ai_manager import get_ai_manager

    ai_manager = get_ai_manager()
    available_models = ai_manager.get_available_models()
    available_model_ids = {m.id for m in available_models}

    # دریافت همه لاگ‌های اخیر
    logs = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id
    ).order_by(desc(ActivityLog.created_at)).limit(500).all()

    fixed_count = 0
    fixed_logs = []

    for log in logs:
        if log.model_id and log.model_id not in available_model_ids and log.model_id != "system":
            # مدل غیرفعال است - پیدا کردن جایگزین
            fallback_model_id = ai_manager.find_fallback_model(log.model_id)

            if fallback_model_id:
                # ذخیره اطلاعات قدیمی
                old_model = log.model_id
                extra = {}
                try:
                    if log.extra_data:
                        extra = json.loads(log.extra_data)
                except:
                    pass

                # ثبت اطلاعات جایگزینی
                extra["model_replacement"] = {
                    "original_model": old_model,
                    "replacement_model": fallback_model_id,
                    "replaced_at": datetime.utcnow().isoformat(),
                    "reason": "مدل اصلی غیرفعال شده - جایگزینی خودکار"
                }

                log.extra_data = json.dumps(extra, ensure_ascii=False)
                fixed_count += 1
                fixed_logs.append({
                    "log_id": log.id,
                    "activity_type": log.activity_type,
                    "original_model": old_model,
                    "replacement_model": fallback_model_id,
                    "created_at": log.created_at.isoformat() if log.created_at else None
                })

                logger.info(f"Fixed disabled model in journal: {old_model} -> {fallback_model_id}")

    db.commit()

    return {
        "success": True,
        "fixed_count": fixed_count,
        "total_checked": len(logs),
        "fixed_logs": fixed_logs,
        "message": f"تعداد {fixed_count} فعالیت با مدل غیرفعال اصلاح شد"
    }


# =====================================
# 🔴 جلوگیری از دور باطل - تابع بررسی تکرار
# =====================================

def check_cycle_prevention(
    db,
    project_id: str,
    activity_type: str,
    field_id: str = None,
    minutes_threshold: int = 5
) -> bool:
    """
    بررسی آیا این فعالیت اخیراً اجرا شده (جلوگیری از دور باطل)

    Returns:
        True اگر فعالیت مجاز است
        False اگر فعالیت تکراری است
    """
    from datetime import timedelta

    threshold_time = datetime.utcnow() - timedelta(minutes=minutes_threshold)

    # بررسی آیا فعالیت مشابه اخیراً اجرا شده
    query = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.activity_type == activity_type,
        ActivityLog.created_at > threshold_time,
        ActivityLog.success == True
    )

    if field_id:
        query = query.filter(ActivityLog.field_id == field_id)

    recent_activity = query.first()

    if recent_activity:
        # فعالیت مشابه در بازه زمانی اخیر وجود دارد
        return False

    return True


# =====================================
# 🆕 خروجی ژورنال - Export
# =====================================

@router.get("/{project_id}/journal/export")
async def export_journal(
    project_id: str,
    format: str = Query("json", description="فرمت خروجی: json, csv, xlsx"),
    days: int = Query(30, ge=1, le=365, description="تعداد روز"),
    activity_type: str = Query(None, description="فیلتر نوع فعالیت"),
    include_operations: bool = Query(False, description="شامل عملیات جزئی"),
    db: Session = Depends(get_db)
):
    """
    خروجی ژورنال فعالیت‌ها در فرمت‌های مختلف

    فرمت‌های پشتیبانی شده:
    - json: فایل JSON با تمام جزئیات
    - csv: فایل CSV برای باز کردن در Excel
    - xlsx: فایل Excel با فرمت‌بندی
    """
    from fastapi.responses import Response
    import csv
    import io

    # بررسی وجود پروژه
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    # دریافت فعالیت‌ها
    since_date = datetime.utcnow() - timedelta(days=days)
    query = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since_date
    )

    if activity_type:
        query = query.filter(ActivityLog.activity_type == activity_type)

    logs = query.order_by(desc(ActivityLog.created_at)).all()

    # تبدیل به دیکشنری
    journal_data = []
    for log in logs:
        log_dict = {
            "id": log.id,
            "activity_type": log.activity_type,
            "model_id": log.model_id,
            "model_provider": log.model_provider,
            "prompt": log.prompt,
            "response": log.response,
            "tokens_used": log.tokens_used,
            "latency_ms": log.latency_ms,
            "success": log.success,
            "error_message": log.error_message,
            "field_id": log.field_id,
            "field_name": log.field_name,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }

        # اضافه کردن عملیات جزئی
        if include_operations:
            operations = db.query(DetailedOperation).filter(
                DetailedOperation.parent_log_id == log.id
            ).order_by(DetailedOperation.sequence_number).all()

            log_dict["operations"] = [
                {
                    "id": op.id,
                    "operation_type": op.operation_type,
                    "summary": op.summary,
                    "target_type": op.target_type,
                    "target_name": op.target_name,
                    "status": op.status,
                    "created_at": op.created_at.isoformat() if op.created_at else None
                }
                for op in operations
            ]

        journal_data.append(log_dict)

    # خروجی JSON
    if format.lower() == "json":
        export_data = {
            "project_id": project_id,
            "project_name": project.name,
            "exported_at": datetime.utcnow().isoformat(),
            "period_days": days,
            "total_activities": len(journal_data),
            "activities": journal_data
        }
        content = json.dumps(export_data, ensure_ascii=False, indent=2)
        return Response(
            content=content,
            media_type="application/json",
            headers={
                "Content-Disposition": f'attachment; filename="journal_{project_id[:8]}_{datetime.now().strftime("%Y%m%d")}.json"'
            }
        )

    # خروجی CSV
    elif format.lower() == "csv":
        output = io.StringIO()
        writer = csv.writer(output)

        # هدر
        headers = [
            "شناسه", "نوع فعالیت", "مدل", "ارائه‌دهنده", "خلاصه",
            "توکن", "تاخیر (ms)", "موفق", "خطا", "فیلد", "تاریخ"
        ]
        writer.writerow(headers)

        # داده‌ها
        for log in journal_data:
            writer.writerow([
                log["id"],
                log["activity_type"],
                log["model_id"],
                log["model_provider"],
                (log["prompt"] or "")[:100],
                log["tokens_used"],
                log["latency_ms"],
                "بله" if log["success"] else "خیر",
                log["error_message"] or "",
                log["field_name"] or "",
                log["created_at"]
            ])

        content = output.getvalue()
        # Add BOM for Excel compatibility with UTF-8
        content = '\ufeff' + content

        return Response(
            content=content.encode('utf-8'),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="journal_{project_id[:8]}_{datetime.now().strftime("%Y%m%d")}.csv"'
            }
        )

    # خروجی XLSX
    elif format.lower() == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="کتابخانه openpyxl نصب نیست. برای خروجی Excel از pip install openpyxl استفاده کنید."
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ژورنال فعالیت‌ها"

        # استایل هدر
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # هدرها
        headers = [
            "شناسه", "نوع فعالیت", "مدل", "ارائه‌دهنده", "خلاصه",
            "توکن", "تاخیر (ms)", "موفق", "خطا", "فیلد", "تاریخ"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # داده‌ها
        for row, log in enumerate(journal_data, 2):
            ws.cell(row=row, column=1, value=log["id"])
            ws.cell(row=row, column=2, value=log["activity_type"])
            ws.cell(row=row, column=3, value=log["model_id"])
            ws.cell(row=row, column=4, value=log["model_provider"])
            ws.cell(row=row, column=5, value=(log["prompt"] or "")[:100])
            ws.cell(row=row, column=6, value=log["tokens_used"])
            ws.cell(row=row, column=7, value=log["latency_ms"])
            ws.cell(row=row, column=8, value="بله" if log["success"] else "خیر")
            ws.cell(row=row, column=9, value=log["error_message"] or "")
            ws.cell(row=row, column=10, value=log["field_name"] or "")
            ws.cell(row=row, column=11, value=log["created_at"])

        # تنظیم عرض ستون‌ها
        column_widths = [15, 15, 12, 12, 40, 10, 12, 8, 30, 20, 22]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # ذخیره در حافظه
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="journal_{project_id[:8]}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
            }
        )

    else:
        raise HTTPException(
            status_code=400,
            detail=f"فرمت '{format}' پشتیبانی نمی‌شود. فرمت‌های مجاز: json, csv, xlsx"
        )


@router.get("/{project_id}/journal/export/stats")
async def export_journal_stats(
    project_id: str,
    days: int = Query(30, ge=1, le=365),
    db: Session = Depends(get_db)
):
    """
    آمار ژورنال برای خروجی
    """
    # بررسی وجود پروژه
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="پروژه یافت نشد")

    since_date = datetime.utcnow() - timedelta(days=days)

    # شمارش کل
    total_count = db.query(ActivityLog).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since_date
    ).count()

    # شمارش بر اساس نوع
    from sqlalchemy import func
    type_counts = db.query(
        ActivityLog.activity_type,
        func.count(ActivityLog.id)
    ).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since_date
    ).group_by(ActivityLog.activity_type).all()

    # شمارش بر اساس مدل
    model_counts = db.query(
        ActivityLog.model_id,
        func.count(ActivityLog.id)
    ).filter(
        ActivityLog.project_id == project_id,
        ActivityLog.created_at >= since_date
    ).group_by(ActivityLog.model_id).all()

    return {
        "success": True,
        "project_id": project_id,
        "project_name": project.name,
        "period_days": days,
        "total_activities": total_count,
        "by_type": {t: c for t, c in type_counts},
        "by_model": {m: c for m, c in model_counts},
        "export_formats": ["json", "csv", "xlsx"]
    }
